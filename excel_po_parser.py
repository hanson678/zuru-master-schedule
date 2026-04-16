# -*- coding: utf-8 -*-
"""Excel PO解析器 — 读取WPS转换后的ZURU PO Excel文件
返回与PDFParser.parse()完全相同的dict格式，实现无缝替换"""
import os, re, logging
import openpyxl


def _normalize_date(s):
    """将各种日期格式统一为YYYY-MM-DD"""
    if not s:
        return ''
    s = str(s).strip().replace('/', '-')
    # YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # DD-MM-YYYY or MM-DD-YYYY
    m = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', s)
    if m:
        a, b, year = int(m.group(1)), int(m.group(2)), m.group(3)
        if a > 12:   # a只能是日，b是月
            return f"{year}-{b:02d}-{a:02d}"
        elif b > 12:  # b只能是日，a是月
            return f"{year}-{a:02d}-{b:02d}"
        else:  # 月日均≤12，ZURU PO固定MM-DD格式
            return f"{year}-{a:02d}-{b:02d}"
    # datetime对象
    from datetime import datetime
    if isinstance(s, datetime):
        return s.strftime('%Y-%m-%d')
    return str(s)


# 国家英文→中文映射
_COUNTRY_MAP = {
    'usa': '美国', 'us': '美国', 'united states': '美国',
    'france': '法国', 'fr': '法国',
    'germany': '德国', 'de': '德国',
    'uk': '英国', 'gb': '英国', 'united kingdom': '英国',
    'australia': '澳大利亚', 'au': '澳大利亚',
    'canada': '加拿大', 'ca': '加拿大',
    'japan': '日本', 'jp': '日本',
    'netherlands': '荷兰', 'nl': '荷兰',
    'spain': '西班牙', 'italy': '意大利',
    'slovakia': '斯洛伐克', 'czech republic': '捷克',
    'poland': '波兰', 'new zealand': '新西兰',
    'south korea': '韩国', 'korea': '韩国',
    'mexico': '墨西哥', 'brazil': '巴西',
    'india': '印度', 'south africa': '南非',
    'china': '中国', 'hong kong': '香港', 'taiwan': '台湾',
    'singapore': '新加坡', 'malaysia': '马来西亚',
    'thailand': '泰国', 'indonesia': '印度尼西亚',
    'russia': '俄罗斯', 'russian fed': '俄罗斯', 'russian federation': '俄罗斯',
    'turkey': '土耳其', 'uae': '阿联酋',
    'sweden': '瑞典', 'norway': '挪威', 'denmark': '丹麦',
    'finland': '芬兰', 'belgium': '比利时', 'austria': '奥地利',
    'switzerland': '瑞士', 'portugal': '葡萄牙',
    'greece': '希腊', 'ireland': '爱尔兰',
    'chile': '智利', 'argentina': '阿根廷',
    'romania': '罗马尼亚', 'hungary': '匈牙利',
    'croatia': '克罗地亚', 'israel': '以色列',
    'guatemala': '危地马拉', 'gt': '危地马拉',
    'uruguay': '乌拉圭', 'uy': '乌拉圭',
    'costa rica': '哥斯达黎加', 'panama': '巴拿马',
    'dominican republic': '多米尼加', 'puerto rico': '波多黎各',
    'ecuador': '厄瓜多尔', 'venezuela': '委内瑞拉',
    'paraguay': '巴拉圭', 'bolivia': '玻利维亚',
    'honduras': '洪都拉斯', 'el salvador': '萨尔瓦多',
    'nicaragua': '尼加拉瓜', 'jamaica': '牙买加',
    'colombia': '哥伦比亚', 'peru': '秘鲁',
    'egypt': '埃及', 'morocco': '摩洛哥',
    'kenya': '肯尼亚', 'nigeria': '尼日利亚',
    'czech': '捷克', 'slovenia': '斯洛文尼亚',
    'slovakia': '斯洛伐克', 'bulgaria': '保加利亚',
    'serbia': '塞尔维亚', 'estonia': '爱沙尼亚',
    'latvia': '拉脱维亚', 'lithuania': '立陶宛',
    'iceland': '冰岛', 'luxembourg': '卢森堡',
    'philippines': '菲律宾', 'vietnam': '越南',
    'cambodia': '柬埔寨', 'myanmar': '缅甸',
    'saudi arabia': '沙特阿拉伯', 'qatar': '卡塔尔',
    'kuwait': '科威特', 'bahrain': '巴林',
    'mongolia': '蒙古', 'mn': '蒙古',
}


def _country_cn(c):
    if not c:
        return ''
    # 去掉换行符后复用PDFParser._country（两套解析器用同一翻译表）
    from pdf_parser import PDFParser as _PP
    c = re.sub(r'[\n\r]+', ' ', str(c))
    return _PP()._country(c)


def _to_float(v):
    """安全转float：处理逗号分隔、空格、字符串等"""
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(re.sub(r'[\s,]+', '', str(v)))
    except (ValueError, TypeError):
        return 0.0


def _to_int(v):
    """安全转int：处理逗号分隔、空格、字符串等"""
    return int(_to_float(v))


class ExcelPOParser:
    """解析WPS转换后的ZURU PO Excel文件"""

    # ===== MPO采购单 (Request for Quotation) 格式 =====

    @staticmethod
    def _is_mpo_rfq(ws):
        """检测是否为MPO采购单格式：R1C1=Purchase Order, R6C4或R6C5以MPO开头"""
        try:
            r1c1 = str(ws.cell(1, 1).value or '').strip()
            r6c1 = str(ws.cell(6, 1).value or '').replace('\n', '').strip()
            if r1c1 != 'Purchase Order' or 'Material Purchase' not in r6c1:
                return False
            r6c5 = str(ws.cell(6, 5).value or '').strip()
            r6c4 = str(ws.cell(6, 4).value or '').strip()
            return r6c5.upper().startswith('MPO') or r6c4.upper().startswith('MPO')
        except:
            return False

    def _parse_mpo_rfq(self, ws):
        """解析MPO采购单：公仔(Line10)+浴巾(Line20)合并为一条双产品entry
        自动适配两种列偏移变体（列偏移A=标准版，列偏移B=整体左移1列）"""
        def _cn_date(v):
            if not v:
                return ''
            s = str(v).replace('\n', '').strip()
            m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', s)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            return _normalize_date(s)

        # 自动检测列偏移：R6C5有MPO号用偏移A，R6C4有MPO号用偏移B
        if str(ws.cell(6, 5).value or '').strip().upper().startswith('MPO'):
            # 偏移A（标准版）：po_date=R5C5, dlv=R5C20, mpo=R6C5, from=R10C5
            # 明细：mat=C4, qty=C8, price=C14, amount=C18, dlv=C11
            _h_col = 5; _dlv_col = 20; _l_mat = 4; _l_qty = 8; _l_price = 14; _l_amt = 18; _l_dlv = 11
        else:
            # 偏移B（整体左移1列）：po_date=R5C4, dlv=R5C19, mpo=R6C4, from=R10C4
            # 明细：mat=C3, qty=C7, price=C13, amount=C17, dlv=C10
            _h_col = 4; _dlv_col = 19; _l_mat = 3; _l_qty = 7; _l_price = 13; _l_amt = 17; _l_dlv = 10

        po_date  = _normalize_date(ws.cell(5, _h_col).value or '')
        delivery = _normalize_date(ws.cell(5, _dlv_col).value or '')
        mpo_no   = str(ws.cell(6, _h_col).value or '').replace('\n', '').strip()
        from_per = str(ws.cell(10, _h_col).value or '').replace('\n', '').strip()
        sku_base = str(ws.cell(13, 2).value or '').replace('\n', '').strip()

        def _read_line(r):
            return {
                'line_no':  str(ws.cell(r, 1).value or '').strip(),
                'qty':      _to_int(ws.cell(r, _l_qty).value),
                'price':    _to_float(ws.cell(r, _l_price).value),
                'amount':   _to_float(ws.cell(r, _l_amt).value),
                'delivery': _cn_date(ws.cell(r, _l_dlv).value),
            }

        line10 = _read_line(13)  # 公仔
        line20 = _read_line(14)  # 浴巾
        dlv = line10['delivery'] or line20['delivery'] or delivery

        # 物料号（货号）：来自明细行的mat_no列，原样写入排期货号列
        mat_no1 = str(ws.cell(13, _l_mat).value or '').replace('\n', '').strip()
        mat_no2 = str(ws.cell(14, _l_mat).value or '').replace('\n', '').strip()

        # 备注：R17C3（固定位置，不随列偏移变化），始终带"Remark: "前缀
        _remark_body = str(ws.cell(17, 3).value or '').replace('\n', ' ').strip()
        remark = f"Remark: {_remark_body}" if _remark_body else "Remark:"

        ln = {
            'line_no':    line10['line_no'],
            'sku':        sku_base,
            'sku_spec':   sku_base,
            'qty':        line10['qty'],    # 公仔数量
            'price':      line10['price'],  # 公仔单价
            'total_usd':  line10['amount'], # 公仔金额
            'qty2':       line20['qty'],    # 浴巾数量
            'price2':     line20['price'],  # 浴巾单价
            'total_usd2': line20['amount'], # 浴巾金额
            'delivery':   dlv,
            'mat_no1':    mat_no1,          # 公仔物料号→排期货号C7
            'mat_no2':    mat_no2,          # 浴巾物料号→排期货号C13
            'remark':     remark,           # 备注→排期备注列
            'is_11962_dual': True,
            'needs_user_confirmation': False,
        }
        header = {
            'po_number': mpo_no, 'po_date': po_date, 'ship_date': dlv,
            'from_person': from_per, 'customer': '', 'destination_cn': '',
            'revision': '',
        }
        return {**header, 'lines': [ln], 'is_cancel': False,
                'raw_text': f'MPO RFQ: {mpo_no}', 'mixed_groups': []}

    # ===== 主入口 =====

    def parse(self, excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        # MPO采购单格式（Request for Quotation）优先检测
        if self._is_mpo_rfq(ws):
            result = self._parse_mpo_rfq(ws)
            wb.close()
            return result

        # 1. 收集所有单元格文本（用于正则匹配头部信息）
        all_text = self._collect_text(ws)

        # 2. 解析头部信息
        header = self._parse_header(all_text)

        # 3. 解析商品行
        lines = self._parse_lines(ws)

        # 3.5 delivery兜底：line无delivery时用header的ship_date
        _fallback_date = header.get('ship_date', '')
        if _fallback_date:
            for ln in lines:
                if not ln.get('delivery'):
                    ln['delivery'] = _fallback_date

        # 3.6 卡板货号合并（SLB/SLD/SLT/SK），必须在混装处理前
        from pdf_parser import PDFParser
        _pdf_parser = PDFParser()
        lines = _pdf_parser._resolve_pallet_groups(lines)
        # 3.7 MEC混装箱处理（与PDF解析器共享逻辑）
        # Excel PO不传all_text（PDF正则不适用于Excel格式，会误匹配卡板行）
        lines = _pdf_parser._resolve_mixed_cartons(lines, '')
        # 混装父行（7154/7153/25257等）：qty改为子行qty之和，原始qty存为carton_count
        # 这些货号的PO结构：父行qty=箱数，STD子行qty=产品件数
        _mixed_parents = {}  # line_no → 父行
        _child_qtys = {}     # line_no → [子行qty列表]
        for ln in lines:
            lno = ln.get('line_no', '')
            if ln.get('is_mixed_carton') and not ln.get('mixed_parent_sku'):
                _mixed_parents[lno] = ln
            elif ln.get('mixed_parent_sku'):
                if lno not in _child_qtys:
                    _child_qtys[lno] = []
                _child_qtys[lno].append(ln.get('qty', 0))
        for lno, parent in _mixed_parents.items():
            cq = _child_qtys.get(lno, [])
            if cq:
                parent['carton_count'] = parent.get('qty', 0)
                parent['qty'] = sum(cq)
                logging.info(f"[混装] {parent.get('sku_spec')} line={lno}: "
                             f"qty={parent['qty']}(子行之和), 箱数={parent['carton_count']}")
        # 过滤掉通用混装子行（有父行的子行不独立入排期）
        # MEC组件行也有mixed_parent_sku但父行不在result中，不能被过滤
        _parent_lnos = set(_mixed_parents.keys())
        lines = [ln for ln in lines
                 if not ln.get('mixed_parent_sku')
                 or ln.get('line_no', '') not in _parent_lnos]
        mixed_groups = getattr(_pdf_parser, '_mixed_groups_info', [])
        # 填充PO号到mixed_groups
        po = header.get('po_number', '')
        for mg in mixed_groups:
            mg['po_number'] = po

        # 同line多行（混装/MEC/多非STD行）→ 主产品行标记needs_user_confirmation
        # _mixed_line_nos：来自_resolve_mixed_cartons的混装检测
        # _multi_row：来自_parse_lines去重前有多个非STD行的line_no（如Fuggler SidekicK+组件行）
        _multi_row = {str(k).strip() for k in getattr(self, '_multi_row_line_nos', set())}
        _mixed_line_nos = {str(mg.get('line_no', '')).strip() for mg in mixed_groups} | _multi_row
        for ln in lines:
            if str(ln.get('line_no', '')).strip() in _mixed_line_nos:
                if not ln.get('is_pallet'):  # 卡板货号已由_resolve_pallet_groups处理，不标混装
                    ln['needs_user_confirmation'] = True

        # 4. 解析备注/包装信息（行迭代法，支持断层空行）
        reqs = self._parse_requirements_from_ws(ws)
        reqs.update(self._parse_requirements(all_text))  # 补充revision字段

        # 5. 检测取消单（单元格文字 + 浮动形状水印）
        is_cancel = self._detect_cancel(all_text) or self._detect_cancel_shape(excel_path)

        wb.close()
        return {**header, 'lines': lines, **reqs,
                'is_cancel': is_cancel, 'raw_text': all_text[:8000],
                'mixed_groups': mixed_groups}

    def _collect_text(self, ws):
        """将所有单元格值拼接为文本（模拟PDF提取文本）
        用制表符分隔单元格，避免与_parse_header正则中的\\s{2,}终止条件冲突
        （如Customer Name后的'COOP FAGHANDEL'被两空格截断为'COOP'）"""
        rows_text = []
        for row in ws.iter_rows(max_col=30, max_row=500):
            cells = []
            for c in row:
                v = c.value
                if v is not None:
                    cells.append(str(v).strip())
            if cells:
                rows_text.append('\t'.join(cells))
        return '\n'.join(rows_text)

    def _parse_header(self, text):
        """用正则从文本中提取头部字段（与PDFParser._header逻辑一致）"""
        def f(p, default=''):
            m = re.search(p, text, re.IGNORECASE)
            return m.group(1).strip() if m else default

        po = (f(r'PO#[:\s]*(4500\d{6})') or
              f(r'Purchase\s+Order[:\s#]*(\d{10})') or
              f(r'PO\s+Number[:\s]*(4\d{9})') or
              f(r'PO[:\s]*#?\s*(4500\d{6})') or
              f(r'Order\s+No\.?[:\s]*(4500\d{6})'))

        dest_raw = (f(r'Destination\s+Country[:\s]*(.+?)(?:\t|\n)') or
                    f(r'Ship\s+To\s+Country[:\s]*(.+?)(?:\t|\n)') or
                    f(r'Destination[:\s]*(.+?)(?:\t|\n)'))
        # 过滤误捕获：若捕到的是字段标签（如"From:"、"Ship:"等），则丢弃
        # 原因：部分PO布局 Destination\tFrom:\tFrance，正则只抓到"From:"就停了
        if dest_raw and re.match(
                r'^(?:From|Ship|Contact|Sales|Loading|Supplier|Buyer|To|Port)\b',
                dest_raw, re.I):
            dest_raw = ''

        ship_date = (f(r'Shipment\s+Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                     f(r'Shipment\s+Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})') or
                     f(r'Ship\s+Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                     f(r'Ship\s+Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})') or
                     f(r'Delivery\s+Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                     f(r'Delivery\s+Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})'))

        # 客户名：用\t或\n截断（单元格边界），而非\s{2,}（客户名内部可能有多空格）
        customer = (f(r'Customer\s+Name[:\s]*(.+?)(?:\t|Payment|Loading|Shipment|\n)') or
                    f(r'Sold\s+To[:\s]*(.+?)(?:\t|Payment|Loading|Shipment|\n)') or
                    f(r'Bill\s+To[:\s]*(.+?)(?:\t|Payment|Loading|Shipment|\n)'))

        po_date_raw = (f(r'(?<![a-zA-Z])Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                       f(r'(?<![a-zA-Z])Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})'))

        cpo = f(r'Customer\s+PO#?[:\s]*(.+?)(?:\t|Loading|Shipment|Payment|\n)')
        if cpo and re.match(r'(?:Loading|Shipment|Payment|Ship|Destination|Supplier)', cpo, re.I):
            cpo = ''

        from_person = f(r'From[:\s]*(.+?)(?:\t|\n)') or f(r'Contact[:\s]*(.+?)(?:\t|\n)')
        # 清理from_person：截断国家名等混入内容
        if from_person:
            from_person = re.split(
                r'\s+(?:Australia|China|United|America|France|Germany|Japan|Korea|'
                r'Guatemala|Uruguay|Costa\s+Rica|Panama|Dominican|Puerto\s+Rico|'
                r'Ecuador|Venezuela|Paraguay|Bolivia|Honduras|El\s+Salvador|'
                r'Nicaragua|Jamaica|Colombia|Peru|Argentina|Chile|'
                r'Sweden|Norway|Denmark|Finland|Belgium|Austria|Switzerland|'
                r'Destination|Sales\s+Order|Loading\s+Port)',
                from_person, maxsplit=1, flags=re.I)[0].strip().rstrip(' ,')

        return {
            'po_number': po,
            'po_date': _normalize_date(po_date_raw),
            'customer': customer,
            'customer_po_header': cpo,
            'from_person': from_person,
            'ship_date': _normalize_date(ship_date),
            'ship_type': f(r'Shipment\s+Type[:\s]*(.+?)(?:\t|\n)'),
            'sales_order': f(r'Sales\s+Order#?[:\s]*(\d+)'),
            'destination': dest_raw,
            # dest_raw为空时，尝试从客户名推断国家（如"ZURU FRANCE" → "France"）
            'destination_cn': _country_cn(dest_raw) or _country_cn(
                re.sub(r'^.*?\b(France|Germany|Spain|Italy|Australia|Japan|Korea|'
                       r'Canada|Mexico|Brazil|Poland|Netherlands|Portugal|'
                       r'Sweden|Norway|Denmark|Finland|Belgium|Austria|Switzerland|'
                       r'United\s+Kingdom|UK|USA|United\s+States)\b.*$',
                       r'\1', customer or '', flags=re.I)
                if customer else ''),
            'loading_port': f(r'Loading\s+Port[:\s]*(.+?)(?:\t|\n)'),
        }

    def _parse_lines(self, ws):
        """从Excel中找到商品行表格并解析"""
        # 1. 找表头行：搜索含"SKU"或"Line"的行
        header_row = None
        col_map = {}  # key: field_name, value: col_index (0-based)

        for row in ws.iter_rows(max_col=100, max_row=100):
            vals = []
            for c in row:
                v = str(c.value).strip().upper() if c.value else ''
                vals.append(v)

            # 检测表头行标志
            has_sku = any('SKU' in v for v in vals)
            has_line = any(v == 'LINE' or v == 'LINE NO' for v in vals)
            has_qty = any('QTY' in v for v in vals)

            if has_sku and (has_qty or has_line):
                header_row = row[0].row
                # 建立列映射
                for i, v in enumerate(vals):
                    vu = v.upper()
                    if vu == 'LINE' or vu == 'LINE NO':
                        col_map['line_no'] = i
                    elif 'SKU' in vu and 'SPEC' not in vu:
                        col_map['sku'] = i
                    elif 'SKU' in vu and 'SPEC' in vu:
                        col_map['sku_spec'] = i
                    elif vu == 'NAME' or '品名' in vu or 'PRODUCT' in vu:
                        col_map['name'] = i
                    elif 'BOM' in vu:
                        col_map['bom'] = i
                    elif 'BARCODE' in vu or 'UPC' in vu or 'EAN' in vu:
                        col_map['barcode'] = i
                    elif 'DELIVERY' in vu or ('DATE' in vu and 'PO' not in vu):
                        col_map['delivery'] = i
                    elif 'PRICE' in vu and 'TOTAL' not in vu and 'STATUS' not in vu:
                        col_map['price'] = i
                    elif 'QTY' in vu and 'PCS' in vu:
                        col_map['qty'] = i
                    elif 'QTY' in vu:
                        col_map.setdefault('qty', i)
                    elif vu.startswith('TOTAL') and 'USD' in vu:
                        col_map['total_usd'] = i
                    elif vu.startswith('TOTAL') and ('CTN' in vu or 'CARTON' in vu):
                        col_map['total_ctns'] = i
                    elif vu.startswith('TOTAL') and 'CBM' in vu:
                        col_map['total_cbm'] = i
                    elif 'CUSTOMER' in vu and 'PO' in vu:
                        col_map['customer_po'] = i
                    elif vu == 'OUTER' or 'OUTER QTY' in vu:
                        col_map['outer_qty'] = i
                    elif vu == 'INNER' or 'INNER QTY' in vu or 'INNER PCS' in vu:
                        col_map['inner_pcs'] = i
                # 自动推断inner_pcs/outer_qty：
                # 先检查上一行(合并单元格标题)来确定PCS列归属
                pcs_indices = [i for i, v in enumerate(vals) if v.upper().strip() == 'PCS']
                if pcs_indices and ('inner_pcs' not in col_map or 'outer_qty' not in col_map):
                    # 读取表头上一行，检查INNER/OUTER合并标题
                    prev_row_vals = []
                    if header_row > 1:
                        try:
                            prev_row = list(ws.iter_rows(min_row=header_row - 1, max_row=header_row - 1, max_col=30))
                            if prev_row:
                                prev_row_vals = [str(c.value or '').strip().upper() for c in prev_row[0]]
                        except:
                            pass
                    for pi in pcs_indices:
                        # 检查上一行同列或相邻列是否有INNER/OUTER标记
                        _label = ''
                        for offset in range(0, min(4, pi + 1)):
                            _ci = pi - offset
                            if _ci >= 0 and _ci < len(prev_row_vals):
                                _pv = prev_row_vals[_ci]
                                if 'INNER' in _pv:
                                    _label = 'inner'
                                    break
                                elif 'OUTER' in _pv:
                                    _label = 'outer'
                                    break
                        if _label == 'inner' and 'inner_pcs' not in col_map:
                            col_map['inner_pcs'] = pi
                        elif _label == 'outer' and 'outer_qty' not in col_map:
                            col_map['outer_qty'] = pi
                    # 兜底：如果上一行没有标记，用位置推断(第一个=inner，第二个=outer)
                    if len(pcs_indices) >= 1 and 'inner_pcs' not in col_map:
                        col_map['inner_pcs'] = pcs_indices[0]
                    if len(pcs_indices) >= 2 and 'outer_qty' not in col_map:
                        col_map['outer_qty'] = pcs_indices[1]
                break

        if not header_row or 'sku' not in col_map:
            return []

        # 2. 从表头下一行开始读取数据
        # 收集行的同时处理续行合并（WPS导出可能把一个单元格拆成多行）
        _line_col = col_map.get('line_no')
        _sku_col = col_map.get('sku')
        _spec_col = col_map.get('sku_spec')

        def _cell_str_with_padding(cell):
            """读取单元格值，对数字值根据number_format恢复前导零"""
            v = cell.value
            if v is None:
                return None
            if isinstance(v, (int, float)):
                s = str(int(v))
                try:
                    nf = cell.number_format or ''
                    # 匹配"000"、"000_"等前导零格式
                    m = re.match(r'^(0+)', nf)
                    if m and len(m.group(1)) > len(s):
                        s = s.zfill(len(m.group(1)))
                except Exception:
                    pass
                return s
            return str(v)

        _merged_rows = []
        # 续行判定辅助列索引
        _price_col = col_map.get('price')
        _qty_col = col_map.get('qty')
        for row in ws.iter_rows(min_row=header_row + 1, max_col=100, max_row=300):
            _cells = list(row)
            vals = [c.value for c in _cells]
            _has_line_no = _line_col is not None and vals[_line_col] is not None and str(vals[_line_col]).strip()
            # 续行判断：无line_no但有SKU或SKU-SPEC部分数据
            _has_partial = False
            if _sku_col is not None and vals[_sku_col] is not None and str(vals[_sku_col]).strip():
                _has_partial = True
            if _spec_col is not None and vals[_spec_col] is not None and str(vals[_spec_col]).strip():
                _has_partial = True
            if not _has_line_no and _has_partial and _merged_rows:
                # 排除STD PRODUCT行（有独立数据，不是续行）
                _name_idx = col_map.get('name')
                _cont_name = str(vals[_name_idx] or '').strip().upper() if _name_idx is not None else ''
                if 'STD' in _cont_name and 'PRODUCT' in _cont_name:
                    _merged_rows.append((_cells, vals))
                    continue
                # 排除有独立price或qty的行（独立数据行，不是续行）
                _cont_has_price = _price_col is not None and vals[_price_col] is not None and _to_float(vals[_price_col]) > 0
                _cont_has_qty = _qty_col is not None and vals[_qty_col] is not None and _to_int(vals[_qty_col]) > 0
                if _cont_has_price or _cont_has_qty:
                    _merged_rows.append((_cells, vals))
                    continue
                # 续行：只拼SKU-SPEC（不拼SKU列，SKU列的续行内容实际是SPEC后半段）
                _prev_cells, _prev_vals = _merged_rows[-1]
                _spec_cv = None
                if _spec_col is not None and _spec_col < len(_cells):
                    _spec_cv = _cell_str_with_padding(_cells[_spec_col])
                    if _spec_cv is not None:
                        _prev_vals[_spec_col] = str(_prev_vals[_spec_col] or '') + _spec_cv
                # SKU列续行内容拼到SPEC：仅当续行SPEC列为空时（说明SPEC溢出到了SKU列）
                # 若续行SPEC列已有值，SKU列的值是SKU自身的续行碎片，不应追加到SPEC
                if _spec_cv is None and _sku_col is not None and _spec_col is not None and _sku_col < len(_cells):
                    _sku_cv = _cell_str_with_padding(_cells[_sku_col])
                    if _sku_cv is not None:
                        _prev_vals[_spec_col] = str(_prev_vals[_spec_col] or '') + _sku_cv
                # Name列也拼接
                _name_col = col_map.get('name')
                if _name_col is not None and _name_col < len(_cells) and vals[_name_col] is not None:
                    _prev_vals[_name_col] = str(_prev_vals[_name_col] or '') + '\n' + str(vals[_name_col])
                # 客户PO列：跨页续行的客PO是当前line的正确值，直接覆盖主行
                _cpo_col = col_map.get('customer_po')
                if _cpo_col is not None and _cpo_col < len(_cells) and vals[_cpo_col] is not None:
                    _cpv = str(vals[_cpo_col]).strip()
                    if _cpv:
                        _prev_vals[_cpo_col] = _cpv
                # 补填delivery/barcode：续行有值且主行为空时继承
                for _fill_key in ('delivery', 'barcode'):
                    _fill_col = col_map.get(_fill_key)
                    if _fill_col is not None and _fill_col < len(_cells) and vals[_fill_col] is not None:
                        _fv = str(vals[_fill_col]).strip()
                        _prev_fv = _prev_vals[_fill_col]
                        if _fv and (_prev_fv is None or str(_prev_fv).strip() == ''):
                            _prev_vals[_fill_col] = vals[_fill_col]
                continue
            _merged_rows.append((_cells, vals))

        lines = []
        for _cells, vals in _merged_rows:
            # 空行或总计行 → 停止
            non_empty = [v for v in vals[:20] if v is not None and str(v).strip()]
            if not non_empty:
                continue
            # 总计行检测：只检查第一个有值的单元格是否以Total开头（避免产品名含TOTAL误判）
            _first_val = str(non_empty[0]).strip().upper()
            if _first_val.startswith('TOTAL'):
                break

            # 读取SKU — 必须有SKU或SKU-SPEC才算有效行
            sku_val = vals[col_map['sku']] if col_map.get('sku') is not None else None
            if not sku_val or not str(sku_val).strip():
                # SKU为空但SPEC有值（跨页续行只拼了SPEC没拼SKU）→ 从SPEC提取SKU
                _spec_idx = col_map.get('sku_spec')
                _spec_val = vals[_spec_idx] if _spec_idx is not None else None
                if _spec_val and str(_spec_val).strip():
                    # SPEC格式如"77770GQ1-S001"，取"-S"前面部分作为SKU
                    _clean_spec = re.sub(r'[\s\n\r]+', '', str(_spec_val).strip())
                    _m = re.match(r'^(.+?)-S\d', _clean_spec)
                    sku_val = _m.group(1) if _m else _clean_spec
                else:
                    continue

            sku = re.sub(r'[\s\n\r]+', '', str(sku_val).strip())

            # 解析各字段
            def get(key, default=''):
                idx = col_map.get(key)
                if idx is None:
                    return default
                v = vals[idx]
                if v is None:
                    return default
                return v

            line_no = str(get('line_no', '')).strip()
            sku_spec = re.sub(r'[\s\n\r]+', '', str(get('sku_spec', sku)).strip())
            name = str(get('name', '')).strip()
            barcode_raw = get('barcode', '')
            barcode = str(int(barcode_raw)) if isinstance(barcode_raw, float) else str(barcode_raw).strip()

            delivery_raw = get('delivery', '')
            from datetime import datetime
            if isinstance(delivery_raw, datetime):
                delivery = delivery_raw.strftime('%Y-%m-%d')
            else:
                delivery = _normalize_date(str(delivery_raw))

            price = _to_float(get('price', 0))
            qty = _to_int(get('qty', 0))
            inner_pcs = _to_int(get('inner_pcs', 0))
            outer_qty = _to_int(get('outer_qty', 0))
            total_usd = _to_float(get('total_usd', 0))
            total_ctns = _to_int(get('total_ctns', 0))
            # outer_qty为0时兜底：从qty÷total_ctns反推，或用inner_pcs
            if outer_qty <= 0 and qty > 0 and total_ctns > 0:
                outer_qty = qty // total_ctns
            if outer_qty <= 0 and inner_pcs > 0:
                outer_qty = inner_pcs
            # qty为0时多级兜底（某些PO无独立QTY列，如196512格式）
            if qty <= 0:
                # 优先：outer_qty × total_ctns（如8箱×1542箱=12336pcs）
                if outer_qty > 0 and total_ctns > 0 and outer_qty * total_ctns > outer_qty:
                    qty = outer_qty * total_ctns
                # 次之：total_usd / price
                elif total_usd > 0 and _to_float(get('price', 0)) > 0:
                    qty = round(total_usd / _to_float(get('price', 0)))
                # 最后：直接用outer或inner
                else:
                    qty = outer_qty or inner_pcs
            # 跳过qty=0的行（空行或汇总行）
            if qty <= 0:
                continue

            # 客PO：保留前导0（Excel可能用数字格式存储带前导0的编号）
            cpo_idx = col_map.get('customer_po')
            cpo = ''
            if cpo_idx is not None and cpo_idx < len(_cells):
                _cpo_str = _cell_str_with_padding(_cells[cpo_idx])
                if _cpo_str is not None:
                    cpo = re.sub(r'[\n\r]+', '', _cpo_str.strip())
            # 清理customer_po：排除CBM等小数误取
            if cpo and re.match(r'^\d+\.\d+$', cpo):
                cpo = ''

            # delivery为空时继承上一行的delivery（MEC复合PO组件行无独立日期）
            if not delivery and lines:
                delivery = lines[-1].get('delivery', '')
            # customer_po为空时继承上一行（同line多行组件行可能无独立客PO）
            if not cpo and lines:
                cpo = lines[-1].get('customer_po', '')
            lines.append({
                'line_no': line_no,
                'sku': sku,
                'item_code': sku_spec if sku_spec != sku else sku,
                'sku_spec': sku_spec if sku_spec != sku else sku,
                'name': name,
                'barcode': barcode,
                'delivery': delivery,
                'price': price,
                'qty': qty,
                'inner_pcs': inner_pcs,
                'outer_qty': outer_qty,
                'total_usd': total_usd,
                'total_ctns': total_ctns,
                'customer_po': cpo,
            })

        # 同一line_no有多行时：STD PRODUCT行保留全部（供混装检测），非STD只保留规格最完整的
        from pdf_parser import PDFParser as _PP
        _PALLET_MAIN_RE = re.compile(r'^(?:\d+(?:SLB|SLD|SLT|SK)\d*|MTQ\d+)(?:-(?!P\d)|$)', re.I)
        _PALLET_PART_RE = re.compile(r'^(?:\d+(?:SLB|SLD|SLT|SK)\d*|MTQ\d+)-P\d', re.I)
        seen_lines = {}   # line_no → 非STD最优行
        std_lines = []    # STD PRODUCT行 + 卡板主行，全部保留
        _non_std_line_count = {}  # 记录每个line_no出现的非STD行数（去重前）
        for ln in lines:
            lno = ln.get('line_no', '')
            if _PP._is_std_product(ln):
                std_lines.append(ln)
                continue
            # 卡板零件行（-P1/-P2等）：丢弃，_resolve_pallet_groups不需要
            _sku_v = str(ln.get('sku_spec') or ln.get('sku') or '')
            if _PALLET_PART_RE.match(_sku_v):
                continue
            # 卡板主行（SLB/SLD/SLT/SK主行）：绕过去重直接保留，供_resolve_pallet_groups识别
            if _PALLET_MAIN_RE.match(_sku_v):
                std_lines.append(ln)
                continue
            _non_std_line_count[lno] = _non_std_line_count.get(lno, 0) + 1
            if lno not in seen_lines:
                seen_lines[lno] = ln
            else:
                prev = seen_lines[lno]
                prev_spec = (prev.get('sku_spec') or '').upper()
                curr_spec = (ln.get('sku_spec') or '').upper()
                prev_has_s = bool(re.search(r'-S\d', prev_spec))
                curr_has_s = bool(re.search(r'-S\d', curr_spec))
                if curr_has_s and not prev_has_s:
                    seen_lines[lno] = ln
                elif curr_has_s and prev_has_s:
                    if ln.get('qty', 0) > prev.get('qty', 0):
                        seen_lines[lno] = ln
        # 保存去重前同line多非STD行的line_no集合（供parse()追加needs_user_confirmation）
        self._multi_row_line_nos = {k for k, v in _non_std_line_count.items() if v > 1 and k}
        # 合并：非STD最优行 + 全部STD行（_resolve_mixed_cartons会处理）
        lines = list(seen_lines.values()) + std_lines

        return lines

    def _parse_requirements_from_ws(self, ws):
        """行迭代法提取备注区域（Tracking Code ~ Order Modifiable Records），支持断层空行。
        比正则更可靠：合并单元格续行不会被漏读。"""
        current_label = ''
        current_parts = []
        sections = {}  # label -> [content lines]

        for row in ws.iter_rows(max_row=600):
            # A列标签
            a_val = str(row[0].value or '').strip() if row else ''

            # 停止：遇到 Order Modifiable Records
            if 'Order Modifiable' in a_val:
                break

            # 内容列：从第2列(B)开始收集所有非空值
            # 跨列内容按行对齐合并（D列存左半、G列存右半时，按\n分行ZIP合并）
            content_parts = []
            for cell in row[1:]:
                if cell.value is not None:
                    v = str(cell.value).strip()
                    if v:
                        content_parts.append(v)
            if not content_parts:
                content = ''
            elif len(content_parts) == 1:
                content = content_parts[0]
            else:
                # 多列都有内容：按\n分行后逐行拼接（保持对齐）
                split_cols = [p.split('\n') for p in content_parts]
                max_lines = max(len(c) for c in split_cols)
                merged_lines = []
                for i in range(max_lines):
                    line_parts = []
                    for col_lines in split_cols:
                        if i < len(col_lines):
                            s = col_lines[i].strip()
                            if s:
                                line_parts.append(s)
                    if line_parts:
                        merged_lines.append(' '.join(line_parts))
                content = '\n'.join(merged_lines)

            # 识别标签
            if 'Tracking Code' in a_val:
                current_label = 'tracking_code'
                current_parts = [content] if content else []
                sections[current_label] = current_parts
            elif 'Packaging Info' in a_val and current_label:
                current_label = 'packaging_info'
                current_parts = [content] if content else []
                sections[current_label] = current_parts
            elif re.match(r'Remark', a_val, re.I) and current_label:
                current_label = 'remark'
                current_parts = [content] if content else []
                sections[current_label] = current_parts
            elif current_label and not a_val and content:
                # 续行：A列为空，内容列有值
                current_parts.append(content)

        tc = '\n'.join(sections.get('tracking_code', []))
        pi = '\n'.join(sections.get('packaging_info', []))
        rm = '\n'.join(sections.get('remark', []))

        # 兜底：客户把Remark内容写到Packaging Info单元格（Alt+Enter换行误操作）
        # 条件：remark为空 且 packaging_info含多行 且 某行行首命中Remark关键字
        # 关键字命中的那行及之后全部切到remark
        if not rm and '\n' in pi:
            _remark_head_kws = ('外箱', 'Remark', 'REMARK', '备注', '注意事项',
                                '注意：', '注意:', '注：', '注:')
            _pi_lines = pi.split('\n')
            _split_idx = -1
            for i, line in enumerate(_pi_lines):
                if i == 0:
                    continue  # 第1行保底是Packaging Info，不参与切分
                if line.lstrip().startswith(_remark_head_kws):
                    _split_idx = i
                    break
            if _split_idx > 0:
                pi = '\n'.join(_pi_lines[:_split_idx]).rstrip()
                rm = '\n'.join(_pi_lines[_split_idx:]).strip()

        return {
            'tracking_code': tc,
            'packaging_info': pi,
            'remark': rm,
        }

    def _parse_requirements(self, text):
        """提取修订记录（revision）。备注/包装信息由 _parse_requirements_from_ws 处理。"""
        revision = ''
        rev_m = re.search(r'Order\s+Modifiable\s+Records\s*(.*?)(?=Special|Additional|Confirmed|$)',
                          text, re.DOTALL | re.I)
        if rev_m:
            entries = []
            for line in rev_m.group(1).strip().split('\n'):
                line = line.strip()
                if not line or ('Revision' in line and '#' in line):
                    continue
                rm = re.match(r'(\d+)\s+(\d{2}-\d{2}-\d{4})\s+(.*)', line)
                if rm:
                    entries.append(f"Rev.{rm.group(1)} ({_normalize_date(rm.group(2))}): {rm.group(3).strip()}")
            if entries:
                revision = '; '.join(entries)
        return {'revision': revision}

    def _detect_cancel(self, text):
        """检测取消单（单元格文字）"""
        clean = re.sub(r'(?:Remark|Packaging\s+Info|备注)[：:\s].*', '', text,
                       flags=re.DOTALL | re.I)
        return '取消' in clean or '取 消' in clean

    def _detect_cancel_shape(self, fp):
        """检测取消单水印：从xlsx ZIP内读浮动形状文字（大字水印不在单元格值中）"""
        import zipfile
        _cancel_kw = ('取消', '取 消', 'CANCEL', 'VOID', 'CANCELLED')
        try:
            with zipfile.ZipFile(fp, 'r') as zf:
                drawing_files = [n for n in zf.namelist() if n.startswith('xl/drawings/') and n.endswith('.xml')]
                for df in drawing_files:
                    xml = zf.read(df).decode('utf-8', errors='ignore')
                    # 提取<a:t>标签内文本
                    texts = re.findall(r'<a:t[^>]*>([^<]+)</a:t>', xml)
                    combined = ''.join(texts)
                    if any(kw in combined for kw in _cancel_kw):
                        return True
        except Exception:
            pass
        return False

    @staticmethod
    def classify_error(filename, error):
        """将解析错误分类为用户能看懂的提示"""
        ext = os.path.splitext(filename)[1].lower() if filename else ''
        err = str(error).lower()
        if ext and ext not in ('.xlsx', '.xls'):
            return {
                'category': 'unsupported_format',
                'title': '文件格式不支持',
                'icon': 'bi-file-earmark-x',
                'color': 'danger',
                'tip': f'上传的是 {ext} 格式文件，系统只支持Excel(.xlsx)格式。'
            }
        return {
            'category': 'parse_failed',
            'title': '解析失败',
            'icon': 'bi-bug',
            'color': 'danger',
            'tip': (f'系统无法从这个文件中读取数据。\n'
                    '可能不是标准ZURU PO格式。\n'
                    '建议手动打开文件查看，手动录入。\n'
                    f'技术详情：{error}')
        }

    @staticmethod
    def validate(data, filename=''):
        """验证解析结果，返回问题/警告列表"""
        issues = []
        lines = data.get('lines', [])
        po = data.get('po_number', '')
        sku_list = ', '.join([ln.get('sku', '?') for ln in lines[:5]]) if lines else ''

        if not po:
            sku_hint = f'\n涉及货号: {sku_list}' if sku_list else ''
            issues.append({
                'category': 'no_po',
                'title': f'未识别到PO号 · {filename}',
                'icon': 'bi-hash',
                'color': 'warning',
                'sku': sku_list,
                'tip': f'文件 {filename} 找不到PO号（正常是4500开头的10位数字）。{sku_hint}'
            })

        if not lines:
            issues.append({
                'category': 'no_lines',
                'title': f'未识别到商品行 · PO {po}' if po else '未识别到商品行',
                'icon': 'bi-list-ul',
                'color': 'danger',
                'tip': f'PO {po} 找不到商品行（Line/SKU/Qty表格）。'
            })

        if not data.get('ship_date') and lines:
            issues.append({
                'category': 'no_ship_date',
                'title': f"缺少出货日期 · PO {po}" if po else '缺少出货日期',
                'icon': 'bi-calendar-x',
                'color': 'warning',
                'sku': sku_list,
                'tip': f"PO {po} 没有检测到出货日期，出货日期列会留空，请手动补上。"
            })

        if not data.get('customer') and lines:
            issues.append({
                'category': 'no_customer',
                'title': f'缺少客户名 · PO {po}' if po else '缺少客户名',
                'icon': 'bi-person-x',
                'color': 'info',
                'sku': sku_list,
                'tip': f'PO {po} 没有识别到客户名(Customer Name)，B列会留空。'
            })

        return issues
