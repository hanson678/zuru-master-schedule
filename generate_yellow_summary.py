# -*- coding: utf-8 -*-
"""扫描总排期整行填充行（不限颜色），按分排期分组，返回JSON供前端展示
判定规则：一行中连续 >=6 个非白非空填充单元格，即视为整行填充
用法：通过Web接口调用 generate_summary(master_path)
"""
import os
import re
import json
import logging
from collections import OrderedDict

logging.basicConfig(level=logging.INFO, format='%(message)s')

# 总排期列号
COL_ITEM = 7   # G列=货号
COL_PO = 4     # D列=PO号
COL_QTY = 9    # I列=数量
COL_CN = 8     # H列=中文名
COL_SHIP = 13  # M列=出货期

# 扫描填充的列范围：总排期完整数据列 A-AB (1-28)
SCAN_COL_START = 1
SCAN_COL_END = 28
# 整行填充判定阈值：连续填充单元格数 >= 此值
MIN_CONTINUOUS_FILLED = 6
# 视为"无填充/白色"的RGB值
_EMPTY_FILL_RGB = {'00000000', 'FFFFFFFF', '00FFFFFF', None, ''}


def _is_colored_fill(cell):
    """判定单元格是否有非白非空填充色"""
    f = cell.fill
    if not f or f.patternType != 'solid':
        return False
    c = f.fgColor
    if not c:
        return False
    rgb = c.rgb
    # theme/indexed 颜色（Excel主题色面板填的）rgb=None但有value
    if rgb is None:
        ctype = getattr(c, 'type', None)
        cval = getattr(c, 'value', None)
        if ctype in ('theme', 'indexed') and cval is not None:
            return True
        return False
    # rgb可能是RGB对象或字符串，安全转str
    try:
        rgb_str = str(rgb).upper()
    except Exception:
        return False
    if rgb_str in _EMPTY_FILL_RGB:
        return False
    return True


def _max_continuous_filled(row, start=SCAN_COL_START, end=SCAN_COL_END):
    """计算一行中最长的连续填充段长度"""
    max_run = 0
    cur_run = 0
    for col_idx in range(start - 1, end):
        if col_idx >= len(row):
            break
        if _is_colored_fill(row[col_idx]):
            cur_run += 1
            if cur_run > max_run:
                max_run = cur_run
        else:
            cur_run = 0
    return max_run


def _get_base(item_str):
    """货号去规格码: 9548UQ1-S001 → 9548UQ1"""
    s = str(item_str).strip()
    m = re.match(r'^(.+?)(-S\d+.*)$', s, re.I)
    return m.group(1).upper() if m else s.upper()


def _load_sub_map():
    """加载分排期映射表"""
    p = os.path.join(os.path.dirname(__file__), 'data', 'sub_schedule_map.json')
    if not os.path.exists(p):
        return {}
    with open(p, 'r', encoding='utf-8') as f:
        return json.load(f)


def _scan_filled_rows(master_path):
    """用openpyxl扫描总排期整行填充行（连续>=MIN_CONTINUOUS_FILLED个非白填充单元格）
    返回行数据列表，不限颜色"""
    import openpyxl
    if not os.path.exists(master_path):
        raise FileNotFoundError(f'总排期文件不存在: {master_path}')

    logging.info(f'正在扫描: {master_path}')
    # read_only=False才能读取单元格填充色
    wb = openpyxl.load_workbook(master_path, data_only=True, read_only=False)
    try:
        # 找总排期sheet（排除旧/取消/汇总）
        ws = None
        for name in wb.sheetnames:
            if '总排期' in name and '旧' not in name and '取消' not in name and '汇总' not in name:
                ws = wb[name]
                break
        if ws is None:
            raise ValueError(f'找不到总排期工作表，可用: {wb.sheetnames}')

        logging.info(f'使用工作表: {ws.title}')

        filled_rows = []
        for row in ws.iter_rows(min_row=2, max_col=SCAN_COL_END, values_only=False):
            # 先判断是否满足"整行填充"：连续>=MIN_CONTINUOUS_FILLED个非白填充
            if _max_continuous_filled(row) < MIN_CONTINUOUS_FILLED:
                continue
            cell_item = row[COL_ITEM - 1]
            item = str(cell_item.value or '').strip()
            if not item:
                continue
            po = str(row[COL_PO - 1].value or '').strip()
            qty = row[COL_QTY - 1].value
            # qty转number防累加出错（Excel里可能是字符串）
            try:
                qty = float(qty) if qty not in (None, '') else 0
            except (TypeError, ValueError):
                qty = 0
            cn_name = str(row[COL_CN - 1].value or '').strip()
            ship = row[COL_SHIP - 1].value
            ship_str = ''
            if ship:
                if hasattr(ship, 'strftime'):
                    ship_str = ship.strftime('%Y-%m-%d')
                else:
                    ship_str = str(ship)
            filled_rows.append({
                'item': item,
                'base': _get_base(item),
                'po': po,
                'qty': qty or 0,
                'cn_name': cn_name,
                'ship_date': ship_str,
            })

        logging.info(f'扫描完成: {len(filled_rows)} 行整行填充')
        return filled_rows
    finally:
        wb.close()


def _group_by_schedule(filled_rows, sub_map):
    """按分排期file→sheet分组"""
    groups = {}
    unmatched = {}

    for yr in filled_rows:
        base = yr['base']
        locs = sub_map.get(base)
        if not locs:
            num_m = re.match(r'^(\d+)', base)
            if num_m:
                locs = sub_map.get(num_m.group(1))
        if locs:
            loc = locs[0]
            key = (loc['file'], loc['sheet'])
            if key not in groups:
                groups[key] = {'items': {}, 'total': 0}
            g = groups[key]
            if base not in g['items']:
                g['items'][base] = {'count': 0, 'cn_name': yr['cn_name'], 'qty_sum': 0}
            g['items'][base]['count'] += 1
            g['items'][base]['qty_sum'] += (yr['qty'] or 0)
            g['total'] += 1
        else:
            if base not in unmatched:
                unmatched[base] = {'count': 0, 'cn_name': yr['cn_name'], 'qty_sum': 0}
            unmatched[base]['count'] += 1
            unmatched[base]['qty_sum'] += (yr['qty'] or 0)

    sorted_groups = OrderedDict(
        sorted(groups.items(), key=lambda x: -x[1]['total'])
    )
    return sorted_groups, unmatched


def generate_summary(master_path):
    """主入口：扫描 → 分组 → 返回JSON数据（不写入任何文件）"""
    sub_map = _load_sub_map()
    if not sub_map:
        raise FileNotFoundError('sub_schedule_map.json 不存在，请先运行 scan_schedules.py')

    filled_rows = _scan_filled_rows(master_path)
    if not filled_rows:
        return {'ok': True, 'total': 0, 'groups': [], 'unmatched': [], 'msg': '无整行填充行'}

    sorted_groups, unmatched = _group_by_schedule(filled_rows, sub_map)

    # 转为前端友好的JSON结构
    groups_list = []
    for (fname, sname), grp in sorted_groups.items():
        items_list = []
        for base, info in sorted(grp['items'].items(), key=lambda x: -x[1]['count']):
            items_list.append({
                'item': base,
                'cn_name': info['cn_name'],
                'count': info['count'],
                'qty_sum': int(info['qty_sum']) if info['qty_sum'] else 0,
            })
        groups_list.append({
            'file': fname,
            'sheet': sname,
            'items': items_list,
            'total': grp['total'],
        })

    unmatched_list = []
    for base, info in sorted(unmatched.items(), key=lambda x: -x[1]['count']):
        unmatched_list.append({
            'item': base,
            'cn_name': info['cn_name'],
            'count': info['count'],
            'qty_sum': int(info['qty_sum']) if info['qty_sum'] else 0,
        })

    grand_total = sum(g['total'] for g in groups_list) + sum(u['count'] for u in unmatched_list)

    return {
        'ok': True,
        'total': grand_total,
        'group_count': len(groups_list),
        'groups': groups_list,
        'unmatched': unmatched_list,
    }


def generate_summary_excel(master_path, output_dir):
    """生成有填充行汇总Excel：按分排期file→sheet分组，每个sheet小计+合计"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    result = generate_summary(master_path)
    if not result.get('ok') or result.get('total', 0) == 0:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '有填充行汇总'

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    subtotal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    unmatch_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['分排期文件', 'Sheet', '货号', '中文名', '行数', '数量']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.freeze_panes = 'A2'

    row = 2
    grand_total_count = 0
    grand_total_qty = 0

    for g in result['groups']:
        group_start = row
        for it in g['items']:
            ws.cell(row=row, column=1, value=g['file']).border = thin_border
            ws.cell(row=row, column=2, value=g['sheet']).border = thin_border
            c3 = ws.cell(row=row, column=3, value=it['item'])
            c3.border = thin_border
            c3.font = Font(name='宋体', size=11, bold=True)
            ws.cell(row=row, column=4, value=it['cn_name']).border = thin_border
            ws.cell(row=row, column=5, value=it['count']).border = thin_border
            ws.cell(row=row, column=6, value=it['qty_sum']).border = thin_border
            row += 1
        # 小计行
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).fill = subtotal_fill
            ws.cell(row=row, column=ci).border = thin_border
            ws.cell(row=row, column=ci).font = Font(name='宋体', size=11, bold=True)
        ws.cell(row=row, column=4, value='小计')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=g['total'])
        grand_total_count += g['total']
        group_qty = sum(it['qty_sum'] for it in g['items'])
        ws.cell(row=row, column=6, value=group_qty)
        grand_total_qty += group_qty
        # 合并文件名和sheet列
        if row - group_start > 1:
            ws.merge_cells(start_row=group_start, start_column=1, end_row=row - 1, end_column=1)
            ws.merge_cells(start_row=group_start, start_column=2, end_row=row - 1, end_column=2)
            ws.cell(row=group_start, column=1).alignment = Alignment(vertical='center', wrap_text=True)
            ws.cell(row=group_start, column=2).alignment = Alignment(vertical='center')
        row += 1

    # 未匹配
    if result['unmatched']:
        ws.cell(row=row, column=1, value='未匹配分排期')
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).fill = unmatch_fill
            ws.cell(row=row, column=ci).border = thin_border
            ws.cell(row=row, column=ci).font = Font(name='宋体', size=11, bold=True)
        row += 1
        unmatch_start = row
        unmatch_count = 0
        unmatch_qty = 0
        for u in result['unmatched']:
            ws.cell(row=row, column=3, value=u['item']).border = thin_border
            ws.cell(row=row, column=3).font = Font(name='宋体', size=11, bold=True)
            ws.cell(row=row, column=4, value=u['cn_name']).border = thin_border
            ws.cell(row=row, column=5, value=u['count']).border = thin_border
            ws.cell(row=row, column=6, value=u['qty_sum']).border = thin_border
            for ci in [1, 2]:
                ws.cell(row=row, column=ci).border = thin_border
            unmatch_count += u['count']
            unmatch_qty += u['qty_sum']
            row += 1
        # 未匹配小计
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).fill = unmatch_fill
            ws.cell(row=row, column=ci).border = thin_border
            ws.cell(row=row, column=ci).font = Font(name='宋体', size=11, bold=True)
        ws.cell(row=row, column=4, value='小计')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=unmatch_count)
        ws.cell(row=row, column=6, value=unmatch_qty)
        grand_total_count += unmatch_count
        grand_total_qty += unmatch_qty
        row += 1

    # 合计行
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = total_fill
        ws.cell(row=row, column=ci).border = thin_border
        ws.cell(row=row, column=ci).font = Font(name='宋体', size=11, bold=True)
    ws.cell(row=row, column=4, value='合计')
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5, value=grand_total_count)
    ws.cell(row=row, column=6, value=grand_total_qty)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(output_dir, exist_ok=True)
    fname = f'有填充行汇总_{ts}.xlsx'
    out_path = os.path.join(output_dir, fname)
    wb.save(out_path)
    wb.close()
    logging.info(f'[汇总Excel] {out_path}')
    return fname
