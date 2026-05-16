# -*- coding: utf-8 -*-
"""总排期写入模块：用WPS COM直接操作Z盘总排期文件，新单追加+修改单定位修改
关键设计：
- openpyxl只读建索引（快速定位已有行），WPS COM写入（不破坏格式）
- 匹配键：PO号+货号+line_no 三元组
- COM打开后检查ReadOnly，被锁则立即退出
- Save前备份.bak防崩溃
"""
import os
import re
import json
import shutil
import logging
from datetime import datetime, timedelta

DEFAULT_MASTER_PATH = r'Z:\各客排期\ZURU生产排期\2025年ZURU总生产排期.xlsx'
SHEET_NAME = '总排期'
BLUE_COM = 15773696  # RGB(0,176,240)

# 总排期列号（与Z盘实际文件一致）
# A=接单期 B=客户 C=走货国 D=PO号 E=客PO F=SKU G=货号# H=中文名
# I=数量 J=内箱 K=外箱 L=总箱 M=出货期 N=验货期
# Y=备注 Z=跟单 AA=单价 AB=金额
COL = {
    'po_date': 1, 'customer': 2, 'dest': 3, 'po': 4, 'cpo': 5,
    'sku_line': 6, 'item': 7, 'cn_name': 8,
    'qty': 9, 'inner': 10, 'outer': 11, 'total_box': 12,
    'ship_date': 13, 'insp_date': 14,
    'remark': 25, 'from_person': 26, 'price': 27, 'amount': 28,
}


def _load_dual_map():
    """加载双排期货号配置"""
    p = os.path.join(os.path.dirname(__file__), 'data', 'dual_schedule_map.json')
    try:
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            return {k.upper(): v for k, v in raw.items() if not k.startswith('_')}
    except (json.JSONDecodeError, OSError) as e:
        logging.error(f'[双排期] 配置文件读取失败: {e}')
    return {}


_DUAL_MAP = _load_dual_map()  # 模块加载时初始化


def _get_dual_map():
    return _DUAL_MAP


def _insert_series(sku_spec, series, mode):
    """在SKU-SPEC中插入系列号，返回新的ITEM#
    mode: append / slt_insert / mid_insert / s_insert / none
    """
    if mode == 'none':
        return sku_spec
    # 防重复：已包含目标系列号则不再追加
    if f'-{series}' in sku_spec.upper().replace(series.upper(), series.upper()):
        parts = sku_spec.upper().split('-')
        if series.upper() in parts:
            return sku_spec
    if mode == 'append':
        # 末尾追加：92123-S001 → 92123-S001-9298
        return f'{sku_spec}-{series}'
    if mode == 'slt_insert':
        # SLT后插入：77721SLT-S001 → 77721SLT-77673-S001
        m = re.match(r'^(.+?SL[TDB])(-.+)$', sku_spec, re.I)
        if m:
            return f'{m.group(1)}-{series}{m.group(2)}'
        return f'{sku_spec}-{series}'
    if mode == 'mid_insert':
        # S00x后插入尾缀前：77785-S001-INT → 77785-S001-77673-INT
        m = re.match(r'^(.+-S\d+)(-.+)?$', sku_spec, re.I)
        if m:
            base = m.group(1)
            suffix = m.group(2) or ''
            return f'{base}-{series}{suffix}'
        return f'{sku_spec}-{series}'
    if mode == 's_insert':
        # S00x前插入（任意前缀任意尾缀，-分隔）：
        #   77869-S001 → 77869-77772-S001
        #   77869-S001-NA → 77869-77772-S001-NA
        #   77869SLT-S001 → 77869SLT-77772-S001
        m = re.match(r'^(.+?)(-S\d+.*)$', sku_spec, re.I)
        if m:
            return f'{m.group(1)}-{series}{m.group(2)}'
        return f'{sku_spec}-{series}'
    return sku_spec


def _expand_dual_items(lines):
    """展开双排期货号：一行变多行，每行ITEM#带不同系列号
    浅拷贝足够：后续只修改sku_spec/sku/_dual_series（均为字符串赋值）"""
    dual_map = _get_dual_map()
    if not dual_map:
        return lines
    result = []
    for ln in lines:
        sku_spec = ln.get('sku_spec', '') or ln.get('sku', '')
        _sku_s = str(sku_spec).strip().upper()
        # 提取数字前缀（如77896）和完整前缀（如MEC426，去掉-S00x及后续）
        base = re.match(r'(\d+)', _sku_s)
        base_num = base.group(1) if base else ''
        full_prefix = re.match(r'([A-Z]*\d+[A-Z]*)', _sku_s)
        full_key = full_prefix.group(1) if full_prefix else ''
        cfg = dual_map.get(base_num) if base_num else None
        if not cfg and full_key:
            cfg = dual_map.get(full_key)
        if not cfg:
            result.append(ln)
            continue
        targets = cfg.get('targets', [])
        mode = cfg.get('mode', 'none')
        for series in targets:
            new_ln = dict(ln)
            new_item = _insert_series(sku_spec, series, mode)
            new_ln['sku_spec'] = new_item
            new_ln['sku'] = new_item
            new_ln['_dual_series'] = series
            # none模式：ITEM#不变但需要在sku_line中加系列号区分索引
            if mode == 'none':
                orig_line_no = ln.get('line_no', '')
                new_ln['line_no'] = f'{orig_line_no}-{series}' if orig_line_no else series
            logging.info(f'[双排期拆行] {sku_spec} → {new_item} (系列{series})')
            result.append(new_ln)
    return result


def _item_base(sku):
    """提取货号基础码：'77772GQ2-S001' → '77772GQ2'"""
    if not sku:
        return ''
    s = re.sub(r'[\s\n]+', '', str(sku).strip()).upper()
    m = re.match(r'(\d+[A-Za-z]*\d*)', s)
    return m.group(1).upper() if m else ''


def _lookup_cn(cn_names, sku):
    """查中文名：先按完整货号精确匹配，再fallback基础码"""
    full = re.sub(r'[\s\n]+', '', str(sku).strip()).upper()
    if full in cn_names:
        return cn_names[full]
    base = _item_base(full)
    return cn_names.get(base, '') if base else ''


def _normalize_po(v):
    """PO号标准化：去掉.0后缀、去空白"""
    s = str(v or '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


# 河源(HY)的Fuggler货号前缀
_HY_PREFIXES = {'15746','15749','15751','15754','15755','15760',
                '15704','15705','15710','15711','15712'}


def _is_fuggler(sku):
    """判断是否Fuggler系列（157开头），验货期-2天"""
    s = re.sub(r'[\s\n]+', '', str(sku or '')).upper()
    base = re.match(r'(\d+)', s)
    if not base:
        return False
    num = base.group(1)
    return num.startswith('157')


def _is_hy(sku):
    """判断是否河源工厂的货号"""
    s = re.sub(r'[\s\n]+', '', str(sku or '')).upper()
    base = re.match(r'(\d{5})', s)
    return base.group(1) in _HY_PREFIXES if base else False


def _calc_inspection(ship_dt, sku=''):
    """验货日期计算
    Fuggler(157开头): 出货-2天; 其他: 出货-4天
    河源(HY): 周六/周日不能验货; 其他: 周日不能验货"""
    if not ship_dt or not hasattr(ship_dt, 'year'):
        return None
    days_before = 2 if _is_fuggler(sku) else 4
    insp_dt = ship_dt - timedelta(days=days_before)
    if _is_hy(sku):
        # 河源：周六→周五，周日→周一
        if insp_dt.weekday() == 5:
            insp_dt -= timedelta(days=1)
        elif insp_dt.weekday() == 6:
            insp_dt += timedelta(days=1)
    else:
        # 非河源：周日→周一
        if insp_dt.weekday() == 6:
            insp_dt += timedelta(days=1)
    return insp_dt


_WEEKDAY_CN = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']


def _fmt_date(dt):
    """日期格式化：2026/5/25周一"""
    if not dt or not hasattr(dt, 'strftime'):
        return ''
    return f"{dt.year}/{dt.month}/{dt.day}{_WEEKDAY_CN[dt.weekday()]}"


def _date_serial(dt):
    """datetime转Excel序列号"""
    if not dt or not hasattr(dt, 'year'):
        return None
    return (dt - datetime(1899, 12, 30)).days


def _build_index(filepath):
    """用openpyxl只读扫描总排期，建立：
    1. (PO+货号+line_no)→行号 的匹配索引
    2. 中文名索引：完整货号→中文名 + 基础码→中文名(最多出现的)
    Returns: index_dict, cn_name_dict, max_row
    """
    import openpyxl
    from collections import Counter
    index = {}       # (po, item_upper, sku_line) → row
    cn_names = {}    # 完整货号/基础码 → cn_name
    _cn_base_counter = {}  # 基础码 → Counter({中文名: 出现次数})
    two_key_count = {}  # (po, item) → 出现次数，用于判断二元组是否唯一
    max_row = 1
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = None
        for sn in wb.sheetnames:
            if sn == SHEET_NAME:
                ws = wb[sn]
                break
        if not ws:
            ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_row=2, max_col=30):
            r = row[0].row
            if r > max_row:
                max_row = r
            po_val = _normalize_po(row[COL['po'] - 1].value)
            item_val = re.sub(r'[\s\n]+', '', str(row[COL['item'] - 1].value or '')).strip().upper()
            sku_line = str(row[COL['sku_line'] - 1].value or '').strip()
            cn_val = str(row[COL['cn_name'] - 1].value or '').strip()

            if po_val and item_val:
                key = (po_val, item_val, sku_line)
                index[key] = r
                key2 = (po_val, item_val)
                two_key_count[key2] = two_key_count.get(key2, 0) + 1
                if key2 not in index:
                    index[key2] = r

            # 中文名索引：按完整货号精确存储 + 基础码统计出现次数
            if item_val and cn_val and any('\u4e00' <= c <= '\u9fff' for c in cn_val):
                cn_names[item_val] = cn_val  # 完整货号精确匹配
                base = _item_base(item_val)
                if base:
                    if base not in _cn_base_counter:
                        _cn_base_counter[base] = Counter()
                    _cn_base_counter[base][cn_val] += 1

        # 基础码取出现次数最多的中文名作为fallback（不覆盖已有的完整货号条目）
        for base, counter in _cn_base_counter.items():
            if base not in cn_names:
                cn_names[base] = counter.most_common(1)[0][0]

        # 同一(PO,货号)出现多行时，删除二元组索引避免错误兜底
        for key2, cnt in two_key_count.items():
            if cnt > 1 and key2 in index:
                del index[key2]

        wb.close()
    except Exception as e:
        logging.warning(f'[总排期索引] 构建失败: {e}')
    return index, cn_names, max_row


def _search_cn_name_com(ws, sku_spec, insert_row, max_row):
    """从总排期G列(货号)筛选同货号行，取H列(中文名)最近的值"""
    target = _item_base(sku_spec)
    if not target:
        return ''
    for sr in range(insert_row - 1, 1, -1):
        h_val = str(ws.Cells(sr, COL['item']).Value or '').strip()
        if h_val and _item_base(h_val) == target:
            i_val = str(ws.Cells(sr, COL['cn_name']).Value or '').strip()
            if i_val and any('\u4e00' <= c <= '\u9fff' for c in i_val):
                return i_val
    for sr in range(insert_row + 1, min(max_row + 1, insert_row + 500)):
        h_val = str(ws.Cells(sr, COL['item']).Value or '').strip()
        if h_val and _item_base(h_val) == target:
            i_val = str(ws.Cells(sr, COL['cn_name']).Value or '').strip()
            if i_val and any('\u4e00' <= c <= '\u9fff' for c in i_val):
                return i_val
    return ''


def write_orders(filepath, orders, export_dir=None, ignored_orders=None):
    """修改单用WPS COM写入总排期，新单生成到独立Excel供复制粘贴

    Returns: {'ok': bool, 'modified': int, 'new_count': int, 'msg': str, 'export_file': str}
    """
    if not os.path.exists(filepath):
        return {'ok': False, 'new': 0, 'modified': 0,
                'msg': f'总排期文件不存在: {filepath}'}

    # 第1步：openpyxl只读建索引（O(1)查找，避免COM逐行扫描）
    logging.info('[总排期] 构建索引...')
    index, cn_names, max_row = _build_index(filepath)
    # 加载中文名直查表兜底（总排期索引中没有的新货号从这里取）
    cn_map_path = os.path.join(os.path.dirname(__file__), 'data', 'item_cn_name_map.json')
    try:
        if os.path.exists(cn_map_path):
            with open(cn_map_path, 'r', encoding='utf-8') as f:
                _raw = json.load(f)
            for k, v in _raw.items():
                if not k.startswith('_') and k.upper() not in cn_names:
                    cn_names[k.upper()] = v.get('cn_name', '') if isinstance(v, dict) else str(v)
    except Exception as e:
        logging.warning(f'[总排期] 中文名直查表加载失败: {e}')
    logging.info(f'[总排期] 索引完成: {len(index)}条记录, {len(cn_names)}个中文名, max_row={max_row}')

    # 第2步：Save前备份
    bak_path = filepath + '.bak'
    try:
        shutil.copy2(filepath, bak_path)
        logging.info(f'[总排期] 备份: {bak_path}')
    except Exception as e:
        logging.warning(f'[总排期] 备份失败: {e}')

    # 第3步：WPS COM打开并写入
    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    wps = None
    wb = None
    try:
        wps = win32com.client.Dispatch('Ket.Application')
        wps.Visible = False
        wps.DisplayAlerts = False
        wb = wps.Workbooks.Open(filepath)

        # 检查ReadOnly — 被锁则立即退出
        if wb.ReadOnly:
            wb.Close(SaveChanges=False)
            wb = None
            return {'ok': False, 'new': 0, 'modified': 0,
                    'msg': '总排期文件被占用（只读模式），请让同事先关闭文件后重试'}

        ws = None
        for i in range(1, wb.Sheets.Count + 1):
            if wb.Sheets(i).Name == SHEET_NAME:
                ws = wb.Sheets(i)
                break
        if not ws:
            ws = wb.Sheets(1)

        new_count = 0
        mod_count = 0
        new_rows = []    # 新单数据收集（不写入Z盘，生成到独立Excel）
        mod_details = [] # 修改单明细
        warnings = []    # 警告信息
        new_details = [] # 新单明细

        for order in orders:
            header = order.get('header') or order
            po = _normalize_po(header.get('po_number', '') or order.get('po_number', ''))
            po_date = header.get('po_date', '') or order.get('po_date', '')
            customer = header.get('customer', '') or order.get('customer', '')
            dest = header.get('destination_cn', '') or order.get('destination_cn', '')
            from_person = header.get('from_person', '') or order.get('from_person', '')
            ship_date_str = header.get('ship_date', '') or order.get('ship_date', '')

            tc = header.get('tracking_code', '') or order.get('tracking_code', '') or ''
            pi = header.get('packaging_info', '') or order.get('packaging_info', '') or ''
            rm = header.get('remark', '') or order.get('remark', '') or ''
            note_parts = []
            if tc: note_parts.append(tc)
            if pi: note_parts.append(f'Packaging Info: {pi}')
            if rm: note_parts.append(f'Remark: {rm}')
            full_note = '\n'.join(note_parts)
            # 检测备注字段缺失（可能是PDF转Excel丢失或图片格式）
            _fname = order.get('filename', '')
            if not pi and not rm:
                warnings.append(f'{_fname}(PO={po}): 包装信息和备注均为空，可能是PDF转Excel时丢失或图片格式，请核对原始PO')
            elif not pi:
                warnings.append(f'{_fname}(PO={po}): 包装信息为空，可能是PDF转Excel时丢失或图片格式，请核对原始PO')
            elif not rm:
                warnings.append(f'{_fname}(PO={po}): 备注(Remark)为空，可能是PDF转Excel时丢失或PO原文未填，请核对原始PO')

            ship_dt = None
            if ship_date_str:
                try:
                    ship_dt = datetime.strptime(str(ship_date_str)[:10], '%Y-%m-%d')
                except:
                    pass

            # 双排期货号展开：一行变多行
            _lines = _expand_dual_items(order.get('lines', []))

            for ln in _lines:
                sku_spec = ln.get('sku_spec', '') or ln.get('sku', '')
                qty = ln.get('qty', 0) or 0
                price = ln.get('price', 0) or 0
                inner_pcs = ln.get('inner_pcs', 0) or 0
                outer_qty = ln.get('outer_qty', 0) or 0
                customer_po = ln.get('customer_po', '')
                is_pallet = ln.get('is_pallet', False)
                pallet_count = ln.get('pallet_count', 0) or 0
                line_no = ln.get('line_no', '')

                # 混装标记
                is_mixed = ln.get('is_mixed_carton', False)
                carton_count = ln.get('carton_count', 0) or 0

                # 外箱：混装=qty÷箱数，卡板=qty÷卡板数，普通=PO的outer_qty
                if is_pallet and pallet_count > 0:
                    if qty > pallet_count:
                        outer_qty = qty // pallet_count
                elif is_mixed and carton_count > 0:
                    outer_qty = qty // carton_count
                    logging.info(f'[混装外箱] {sku_spec}: {qty}//{carton_count}={outer_qty}')

                # 总箱数：卡板=卡板数，混装=箱数，普通=qty÷外箱
                if is_pallet and pallet_count > 0:
                    total_ctns = pallet_count
                elif is_mixed and carton_count > 0:
                    total_ctns = carton_count
                    logging.info(f'[混装总箱] {sku_spec}: 箱数={carton_count}')
                elif outer_qty > 0:
                    total_ctns = qty // outer_qty if qty else 0
                else:
                    total_ctns = ln.get('total_ctns', 0) or 0

                # 金额：卡板/混装=总箱×单价，普通=数量×单价
                if is_pallet and pallet_count > 0:
                    total_usd = total_ctns * price
                elif is_mixed and carton_count > 0:
                    total_usd = total_ctns * price
                    logging.info(f'[混装金额] {sku_spec}: {total_ctns}*{price}={total_usd}')
                else:
                    total_usd = qty * price

                line_ship = ln.get('delivery', '') or ship_date_str
                line_ship_dt = None
                if line_ship:
                    try:
                        line_ship_dt = datetime.strptime(str(line_ship)[:10], '%Y-%m-%d')
                    except:
                        line_ship_dt = ship_dt
                else:
                    line_ship_dt = ship_dt
                insp_dt = _calc_inspection(line_ship_dt, sku_spec)

                f_sku = f"{po}-{line_no}" if po and line_no else ''
                item_upper = re.sub(r'[\s\n]+', '', str(sku_spec)).strip().upper()

                # 索引查找：先三元组，再二元组，最后截取到-S00x兜底
                existing_row = index.get((po, item_upper, f_sku))
                if not existing_row:
                    existing_row = index.get((po, item_upper))
                if not existing_row:
                    # 第三层兜底：货号可能粘连了产品描述（如77889SLT-S001-PRODBC）
                    # 截取到 -S00x 部分再匹配
                    _trunc = re.match(r'(.+-S\d+)', item_upper)
                    if _trunc:
                        _item_trunc = _trunc.group(1)
                        existing_row = index.get((po, _item_trunc))
                        if existing_row:
                            logging.info(f'[索引匹配] 基础码兜底: {item_upper} -> {_item_trunc}')
                logging.info(f'[索引匹配] po={po} item={item_upper} f_sku={f_sku} '
                             f'-> row={existing_row or "未找到(新单)"}')

                if existing_row:
                    # ========== 修改单 ==========
                    r = existing_row
                    updates = [(COL['qty'], qty)]
                    if inner_pcs:
                        updates.append((COL['inner'], inner_pcs))
                    if outer_qty:
                        updates.append((COL['outer'], outer_qty))
                    # 总箱：公式=数量÷外箱（外箱=0时留空避免#DIV/0!）
                    updates.append((COL['total_box'], f'=IF(RC{COL["outer"]}=0,"",RC{COL["qty"]}/RC{COL["outer"]})', True))
                    if customer_po:
                        updates.append((COL['cpo'], customer_po))
                    if line_ship_dt:
                        updates.append((COL['ship_date'], _date_serial(line_ship_dt)))
                    if insp_dt:
                        updates.append((COL['insp_date'], _date_serial(insp_dt)))
                    if price:
                        updates.append((COL['price'], round(price, 4)))
                    # 金额：卡板=总箱×单价，普通=数量×单价
                    if is_pallet and pallet_count > 0:
                        updates.append((COL['amount'], f'=RC{COL["total_box"]}*RC{COL["price"]}', True))
                    else:
                        updates.append((COL['amount'], f'=RC{COL["qty"]}*RC{COL["price"]}', True))

                    # 列号→中文名映射（用于变化明细展示）
                    _COL_NAMES = {
                        COL['qty']: '数量', COL['inner']: '内箱', COL['outer']: '外箱',
                        COL['total_box']: '总箱', COL['cpo']: '客PO',
                        COL['ship_date']: '出货期', COL['insp_date']: '验货期',
                        COL['price']: '单价', COL['amount']: '金额',
                    }
                    changed_fields = []  # 记录具体变化
                    for _upd in updates:
                        is_formula = len(_upd) == 3 and _upd[2] is True
                        col_num, new_val = _upd[0], _upd[1]
                        old_val = ws.Cells(r, col_num).Value
                        # 公式列：直接写入不比较（公式结果依赖其他单元格）
                        if is_formula:
                            old_show = str(old_val or '')
                            ws.Cells(r, col_num).FormulaR1C1 = new_val
                            ws.Cells(r, col_num).Interior.Color = BLUE_COM
                            col_label = _COL_NAMES.get(col_num, f'列{col_num}')
                            new_display = '公式'
                            changed_fields.append(f'{col_label} {old_show}→{new_display}')
                            continue
                        # COM返回datetime时转为序列号比较
                        old_display = old_val  # 用于展示的原始值
                        if hasattr(old_val, 'year'):
                            try:
                                old_display = _fmt_date(old_val.replace(tzinfo=None))
                                old_val = (old_val.replace(tzinfo=None) - datetime(1899, 12, 30)).days
                            except Exception:
                                pass
                        # 值相同则跳过（不标蓝）
                        try:
                            if old_val is not None and new_val is not None:
                                o_f = float(old_val)
                                n_f = float(new_val)
                                if abs(o_f - n_f) < 0.0001:
                                    continue
                        except (ValueError, TypeError):
                            pass
                        o_s = str(old_val or '').strip()
                        n_s = str(new_val or '').strip()
                        if o_s.endswith('.0'):
                            o_s = o_s[:-2]
                        if n_s.endswith('.0'):
                            n_s = n_s[:-2]
                        if o_s == n_s:
                            continue
                        # PO号/客PO：字符串写入，保留前导零（如客PO 0581402）
                        if col_num in (COL['po'], COL['cpo']):
                            _pv = str(new_val).strip()
                            if _pv.endswith('.0'):
                                _pv = _pv[:-2]
                            ws.Cells(r, col_num).NumberFormat = '@'
                            ws.Cells(r, col_num).Value = _pv
                        else:
                            ws.Cells(r, col_num).Value = new_val
                        ws.Cells(r, col_num).Interior.Color = BLUE_COM
                        # 记录变化：日期列用可读格式
                        col_label = _COL_NAMES.get(col_num, f'列{col_num}')
                        if col_num in (COL['ship_date'], COL['insp_date']):
                            new_display = _fmt_date(datetime(1899, 12, 30) + timedelta(days=int(new_val))) if new_val else ''
                            old_show = str(old_display or '')
                        else:
                            old_show = o_s
                            new_display = n_s
                        changed_fields.append(f'{col_label} {old_show}→{new_display}')

                    if changed_fields:
                        mod_count += 1
                        mod_details.append({
                            'item': sku_spec, 'row': r, 'po': po,
                            'changes': changed_fields
                        })
                        logging.info(f'[总排期修改] row={r} PO={po} SKU={sku_spec} 变化: {", ".join(changed_fields)}')
                else:
                    # ========== 新单 → 收集到列表（不写入Z盘）==========
                    cn_name = _lookup_cn(cn_names, sku_spec)
                    # 接单日期转datetime
                    po_date_dt = None
                    if po_date:
                        try:
                            po_date_dt = datetime.strptime(str(po_date)[:10], '%Y-%m-%d')
                        except Exception:
                            pass
                    new_rows.append({
                        'po_date': po_date_dt, 'customer': customer, 'dest': dest,
                        'po': po, 'cpo': customer_po, 'sku_line': f_sku,
                        'item': sku_spec, 'cn_name': cn_name,
                        'qty': qty, 'inner': inner_pcs, 'outer': outer_qty,
                        'total_box': '__FORMULA_TOTAL_BOX__',
                        'ship_date': line_ship_dt,
                        'insp_date': insp_dt,
                        'remark': full_note,
                        'from_person': from_person.strip() if from_person else '',
                        'price': round(price, 4) if price else '',
                        'amount': '__FORMULA_AMOUNT_PALLET__' if (is_pallet and pallet_count > 0) else '__FORMULA_AMOUNT__',
                    })
                    new_count += 1
                    new_details.append(sku_spec)
                    logging.info(f'[总排期新单] PO={po} SKU={sku_spec} (收集到Excel)')

        if mod_count:
            wb.Save()
            logging.info(f'[总排期] 修改{mod_count}行已保存')

        # 组装黑名单行数据（与new_rows相同格式）
        ignored_rows = []
        if ignored_orders:
            for order in ignored_orders:
                header = order.get('header') or order
                _po = _normalize_po(header.get('po_number', '') or order.get('po_number', ''))
                _po_date = header.get('po_date', '') or order.get('po_date', '')
                _customer = header.get('customer', '') or order.get('customer', '')
                _dest = header.get('destination_cn', '') or order.get('destination_cn', '')
                _from = header.get('from_person', '') or order.get('from_person', '')
                for ln in order.get('lines', []):
                    _sku = ln.get('sku_spec', '') or ln.get('sku', '')
                    _qty = ln.get('qty', 0) or 0
                    _price = ln.get('price', 0) or 0
                    _inner = ln.get('inner_pcs', 0) or 0
                    _outer = ln.get('outer_qty', 0) or 0
                    _cpo = ln.get('customer_po', '')
                    _line_no = ln.get('line_no', '')
                    _f_sku = f"{_po}-{_line_no}" if _po and _line_no else ''
                    _ship = ln.get('delivery', '') or header.get('ship_date', '') or order.get('ship_date', '')
                    _ship_dt = None
                    if _ship:
                        try:
                            _ship_dt = datetime.strptime(str(_ship)[:10], '%Y-%m-%d')
                        except Exception:
                            pass
                    _po_date_dt = None
                    if _po_date:
                        try:
                            _po_date_dt = datetime.strptime(str(_po_date)[:10], '%Y-%m-%d')
                        except Exception:
                            pass
                    ignored_rows.append({
                        'po_date': _po_date_dt, 'customer': _customer, 'dest': _dest,
                        'po': _po, 'cpo': _cpo, 'sku_line': _f_sku,
                        'item': _sku, 'cn_name': _lookup_cn(cn_names, _sku),
                        'qty': _qty, 'inner': _inner, 'outer': _outer,
                        'total_box': '__FORMULA_TOTAL_BOX__',
                        'ship_date': _ship_dt, 'insp_date': _calc_inspection(_ship_dt, _sku),
                        'remark': '',
                        'from_person': _from.strip() if _from else '',
                        'price': round(_price, 4) if _price else '',
                        'amount': '__FORMULA_AMOUNT__',
                    })

        # 新单生成到独立Excel
        export_file = ''
        if (new_rows or ignored_rows) and export_dir:
            export_file = _generate_new_rows_excel(new_rows, export_dir,
                                                   ignored_rows=ignored_rows)

        parts = []
        if mod_count: parts.append(f'修改{mod_count}行(已写入总排期)')
        if new_count: parts.append(f'新增{new_count}行(已生成Excel)')
        msg = '、'.join(parts) if parts else '无变化'
        logging.info(f'[总排期] {msg}')
        return {'ok': True, 'modified': mod_count, 'new_count': new_count,
                'msg': msg, 'export_file': export_file,
                'mod_details': mod_details, 'new_details': new_details,
                'warnings': warnings}

    except Exception as e:
        logging.error(f'[总排期] 写入失败: {e}')
        raise
    finally:
        try:
            if wb:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if wps:
                wps.Quit()
        except:
            pass
        pythoncom.CoUninitialize()


def _generate_new_rows_excel(new_rows, output_dir, ignored_rows=None):
    """把新单数据生成为格式化Excel，按分排期sheet分组
    第一个sheet"新单数据"放全部行，后续sheet按分排期sheet名分组
    ignored_rows不为空时额外生成"黑名单"sheet"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    col_order = [
        ('po_date', '接单期'), ('customer', '第三方客户名称'), ('dest', '走货国家'),
        ('po', 'PO号'), ('cpo', '客PO'), ('sku_line', '-110'),
        ('item', '货号#'), ('cn_name', '中文名'),
        ('qty', 'PO数量(pcs)'), ('inner', '内箱装箱数量(pcs)'),
        ('outer', '外箱装箱数量(只)'), ('total_box', '总箱数'),
        ('ship_date', 'ZURU订单走货日期'), ('insp_date', 'ZURU验货日期'),
        ('remark', '备注'), ('from_person', 'ZURU跟单员'),
        ('price', '单价USD'), ('amount', '金额USD'),
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    _DATE_KEYS = {'po_date', 'ship_date', 'insp_date'}
    _key_to_col = {}
    for ci, (key, _) in enumerate(col_order, start=1):
        _key_to_col[key] = openpyxl.utils.get_column_letter(ci)
    _QTY_COL = _key_to_col.get('qty', 'I')
    _OUTER_COL = _key_to_col.get('outer', 'K')
    _TOTAL_COL = _key_to_col.get('total_box', 'L')
    _PRICE_COL = _key_to_col.get('price', 'AA')

    col_widths = {
        'po_date': 18, 'customer': 18, 'dest': 8, 'po': 14, 'cpo': 14,
        'sku_line': 18, 'item': 22, 'cn_name': 18, 'qty': 10, 'inner': 8,
        'outer': 8, 'total_box': 8, 'ship_date': 18, 'insp_date': 18,
        'remark': 30, 'from_person': 12, 'price': 10, 'amount': 12,
    }

    def _write_header(ws):
        for ci, (_, label) in enumerate(col_order, start=1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        for ci, (key, _) in enumerate(col_order, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths.get(key, 12)
        ws.freeze_panes = 'A2'

    def _write_rows(ws, rows, start_row=2):
        for ri_offset, row_data in enumerate(rows):
            ri = start_row + ri_offset
            for ci, (key, _) in enumerate(col_order, start=1):
                val = row_data.get(key, '')
                if val == '__FORMULA_TOTAL_BOX__':
                    val = f'=IF({_OUTER_COL}{ri}=0,"",{_QTY_COL}{ri}/{_OUTER_COL}{ri})'
                elif val == '__FORMULA_AMOUNT__':
                    val = f'={_QTY_COL}{ri}*{_PRICE_COL}{ri}'
                elif val == '__FORMULA_AMOUNT_PALLET__':
                    val = f'={_TOTAL_COL}{ri}*{_PRICE_COL}{ri}'
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(name='宋体', size=11, color='000000')
                cell.fill = blue_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=(key == 'remark'))
                if key in _DATE_KEYS and val:
                    cell.number_format = 'yyyy/m/d'
                elif key in ('po', 'cpo') and val:
                    # 字符串写入，保留前导零（如客PO 0581402）
                    _pv = str(val).strip()
                    if _pv.endswith('.0'):
                        _pv = _pv[:-2]
                    cell.value = _pv
                    cell.number_format = '@'

    wb = openpyxl.Workbook()

    # ── Sheet1: 全部新单（保留原有行为） ──
    ws_all = wb.active
    ws_all.title = '新单数据'
    _write_header(ws_all)
    _write_rows(ws_all, new_rows)

    # ── 按分排期sheet分组，创建额外sheet ──
    try:
        sub_map = _load_sub_schedule_map()
        if sub_map:
            # 给每行打标：归属哪个分排期sheet
            grouped = {}  # {sheet_name: [row_data]}
            no_match = []
            for row_data in new_rows:
                item = str(row_data.get('item', '')).strip()
                base = re.match(r'^(.+?)(-S\d+.*)$', item, re.I)
                base = base.group(1).upper() if base else item.upper()
                # 查映射
                locs = _find_in_sub_map(sub_map, base)
                if locs:
                    sname = _best_sheet_name(locs, base)
                    if sname not in grouped:
                        grouped[sname] = []
                    grouped[sname].append(row_data)
                else:
                    no_match.append(row_data)

            # 按行数降序创建sheet
            for sname in sorted(grouped.keys(), key=lambda k: -len(grouped[k])):
                rows = grouped[sname]
                # sheet名最长31字符，去非法字符
                safe_name = re.sub(r'[\[\]:*?/\\]', '', sname)[:31]
                # 避免与已有sheet重名
                existing = {s.title for s in wb.worksheets}
                if safe_name in existing:
                    safe_name = safe_name[:28] + '_' + str(len(existing))
                ws = wb.create_sheet(title=safe_name)
                _write_header(ws)
                _write_rows(ws, rows)

            # 未匹配的放"其他"sheet
            if no_match:
                ws_other = wb.create_sheet(title='未匹配')
                _write_header(ws_other)
                _write_rows(ws_other, no_match)
    except Exception as e:
        logging.warning(f'[新单Excel] 分排期分组失败（不影响主sheet）: {e}')

    # ── 黑名单sheet：被忽略的货号行 ──
    if ignored_rows:
        try:
            ws_ig = wb.create_sheet(title='黑名单')
            _write_header(ws_ig)
            _write_rows(ws_ig, ignored_rows)
            logging.info(f'[新单Excel] 黑名单sheet: {len(ignored_rows)}行')
        except Exception as e:
            logging.warning(f'[新单Excel] 黑名单sheet生成失败: {e}')

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(output_dir, exist_ok=True)
    fname = f'总排期新单_{ts}.xlsx'
    out_path = os.path.join(output_dir, fname)
    sheet_count = len(wb.worksheets)
    wb.save(out_path)
    wb.close()
    logging.info(f'[新单Excel] {out_path}，{len(new_rows)}行，{sheet_count}个sheet')
    return fname


def generate_excel(orders, output_dir):
    """把PO数据生成为格式化Excel表格（不写入Z盘总排期，供用户手动粘贴）
    列顺序与总排期一致，蓝色填充标记新数据
    Returns: {'ok': True, 'path': 文件路径, 'count': 行数, 'msg': 描述}
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 表头（与总排期COL对应，无系统货号列）
    col_order = [
        ('po_date', '接单期'), ('customer', '第三方客户名称'), ('dest', '走货国家'),
        ('po', 'PO号'), ('cpo', '客PO'), ('sku_line', 'SKU'),
        ('item', '货号#'), ('cn_name', '中文名'),
        ('qty', 'PO数量(pcs)'), ('inner', '内箱装箱数量(pcs)'),
        ('outer', '外箱装箱数量(pcs)'), ('total_box', '总箱数'),
        ('ship_date', 'ZURU订单走货日期'), ('insp_date', 'ZURU验货日期'),
        ('remark', '备注'), ('from_person', 'ZURU跟单员'),
        ('price', '单价USD'), ('amount', '金额USD'),
    ]

    # 加载中文名直查表
    cn_map = {}
    cn_path = os.path.join(os.path.dirname(__file__), 'data', 'item_cn_name_map.json')
    try:
        if os.path.exists(cn_path):
            with open(cn_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            cn_map = {k.upper(): v.get('cn_name', '') if isinstance(v, dict) else str(v)
                      for k, v in raw.items() if not k.startswith('_')}
    except Exception:
        pass

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    data_font = Font(name='宋体', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '总排期数据'

    # 写表头
    for ci, (key, label) in enumerate(col_order, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    row_idx = 2
    for order in orders:
        header = order.get('header') or order
        po = _normalize_po(header.get('po_number', '') or order.get('po_number', ''))
        po_date = header.get('po_date', '') or order.get('po_date', '')
        customer = header.get('customer', '') or order.get('customer', '')
        dest = header.get('destination_cn', '') or order.get('destination_cn', '')
        from_person = header.get('from_person', '') or order.get('from_person', '')
        ship_date_str = header.get('ship_date', '') or order.get('ship_date', '')

        # 备注
        tc = header.get('tracking_code', '') or order.get('tracking_code', '') or ''
        pi = header.get('packaging_info', '') or order.get('packaging_info', '') or ''
        rm = header.get('remark', '') or order.get('remark', '') or ''
        note_parts = []
        if tc: note_parts.append(tc)
        if pi: note_parts.append(f'Packaging Info: {pi}')
        if rm: note_parts.append(f'Remark: {rm}')
        full_note = '\n'.join(note_parts)

        ship_dt = None
        if ship_date_str:
            try:
                ship_dt = datetime.strptime(str(ship_date_str)[:10], '%Y-%m-%d')
            except Exception:
                pass

        # 双排期展开
        _lines = _expand_dual_items(order.get('lines', []))

        for ln in _lines:
            sku_spec = ln.get('sku_spec', '') or ln.get('sku', '')
            qty = ln.get('qty', 0) or 0
            price = ln.get('price', 0) or 0
            inner_pcs = ln.get('inner_pcs', 0) or 0
            outer_qty = ln.get('outer_qty', 0) or 0
            customer_po = ln.get('customer_po', '')
            is_pallet = ln.get('is_pallet', False)
            pallet_count = ln.get('pallet_count', 0) or 0
            line_no = ln.get('line_no', '')
            is_mixed = ln.get('is_mixed_carton', False)
            carton_count = ln.get('carton_count', 0) or 0

            # 外箱
            if is_pallet and pallet_count > 0:
                if qty > pallet_count:
                    outer_qty = qty // pallet_count
            elif is_mixed and carton_count > 0:
                outer_qty = qty // carton_count

            # 总箱
            if is_pallet and pallet_count > 0:
                total_ctns = pallet_count
            elif is_mixed and carton_count > 0:
                total_ctns = carton_count
            elif outer_qty > 0:
                total_ctns = qty // outer_qty if qty else 0
            else:
                total_ctns = ln.get('total_ctns', 0) or 0

            # 金额
            if (is_pallet and pallet_count > 0) or (is_mixed and carton_count > 0):
                total_usd = total_ctns * price
            else:
                total_usd = qty * price

            # 日期
            line_ship = ln.get('delivery', '') or ship_date_str
            line_ship_dt = None
            if line_ship:
                try:
                    line_ship_dt = datetime.strptime(str(line_ship)[:10], '%Y-%m-%d')
                except Exception:
                    line_ship_dt = ship_dt
            else:
                line_ship_dt = ship_dt
            insp_dt = _calc_inspection(line_ship_dt, sku_spec)

            f_sku = f"{po}-{line_no}" if po and line_no else ''
            cn_name = _lookup_cn(cn_map, sku_spec)

            # 接单日期转datetime
            po_date_dt = None
            if po_date:
                try:
                    po_date_dt = datetime.strptime(str(po_date)[:10], '%Y-%m-%d')
                except Exception:
                    pass

            # 组装行数据
            row_data = {
                'po_date': po_date_dt,
                'customer': customer,
                'dest': dest,
                'po': po,
                'cpo': customer_po,
                'sku_line': f_sku,

                'item': sku_spec,
                'cn_name': cn_name,
                'qty': qty,
                'inner': inner_pcs,
                'outer': outer_qty,
                'total_box': '__FORMULA_TOTAL_BOX__',
                'ship_date': line_ship_dt,
                'insp_date': insp_dt,
                'remark': full_note,
                'from_person': from_person.strip() if from_person else '',
                'price': round(price, 4) if price else '',
                'amount': '__FORMULA_AMOUNT_PALLET__' if (is_pallet and pallet_count > 0) else '__FORMULA_AMOUNT__',
            }

            for ci, (key, _) in enumerate(col_order, start=1):
                val = row_data.get(key, '')
                # 公式标记替换
                if val == '__FORMULA_TOTAL_BOX__':
                    _qc = openpyxl.utils.get_column_letter({k: i+1 for i, (k,_) in enumerate(col_order)}['qty'])
                    _oc = openpyxl.utils.get_column_letter({k: i+1 for i, (k,_) in enumerate(col_order)}['outer'])
                    val = f'=IF({_oc}{row_idx}=0,"",{_qc}{row_idx}/{_oc}{row_idx})'
                elif val == '__FORMULA_AMOUNT__':
                    _qc = openpyxl.utils.get_column_letter({k: i+1 for i, (k,_) in enumerate(col_order)}['qty'])
                    _pc = openpyxl.utils.get_column_letter({k: i+1 for i, (k,_) in enumerate(col_order)}['price'])
                    val = f'={_qc}{row_idx}*{_pc}{row_idx}'
                elif val == '__FORMULA_AMOUNT_PALLET__':
                    _tc = openpyxl.utils.get_column_letter({k: i+1 for i, (k,_) in enumerate(col_order)}['total_box'])
                    _pc = openpyxl.utils.get_column_letter({k: i+1 for i, (k,_) in enumerate(col_order)}['price'])
                    val = f'={_tc}{row_idx}*{_pc}{row_idx}'
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font = Font(name='宋体', size=11, color='000000')
                cell.fill = blue_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=(key == 'remark'))
                if key in ('po_date', 'ship_date', 'insp_date') and val:
                    cell.number_format = 'yyyy/m/d'
                elif key in ('po', 'cpo') and val:
                    _pv = str(val).strip()
                    if _pv.endswith('.0'):
                        _pv = _pv[:-2]
                    cell.value = _pv
                    cell.number_format = '@'
            row_idx += 1

    # 列宽
    col_widths = {
        'po_date': 18, 'customer': 18, 'dest': 8, 'po': 14, 'cpo': 14,
        'sku_line': 18, 'item': 22, 'cn_name': 18,
        'qty': 10, 'inner': 8, 'outer': 8, 'total_box': 8,
        'ship_date': 18, 'insp_date': 18, 'remark': 30, 'from_person': 12,
        'price': 10, 'amount': 12,
    }
    for ci, (key, _) in enumerate(col_order, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths.get(key, 12)

    # 冻结首行
    ws.freeze_panes = 'A2'

    total_rows = row_idx - 2
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f'总排期数据_{ts}.xlsx')
    wb.save(out_path)
    wb.close()
    logging.info(f'[生成Excel] {out_path}，{total_rows}行')
    return {'ok': True, 'path': out_path, 'count': total_rows,
            'msg': f'已生成{total_rows}行数据'}


# ── 分排期归属查找（独立于write_orders，纯只读） ──────────────────

_sub_map_cache = {}  # 缓存，避免每次重新加载

def _find_in_sub_map(sub_map, base):
    """在sub_schedule_map中查找货号，返回locs列表或None
    查找优先级：1)精确匹配 2)数字前缀 3)base是某个key的前缀（如77896SLT匹配77896SLT-77772）
    """
    # 1) 精确匹配
    if base in sub_map:
        return sub_map[base]
    # 2) 数字前缀
    num_m = re.match(r'^(\d+)', base)
    if num_m:
        prefix = num_m.group(1)
        if prefix in sub_map:
            return sub_map[prefix]
    # 3) base是某个key的前缀（如77896SLT → 77896SLT-77772）
    for k, v in sub_map.items():
        if k.startswith(base + '-') or k.startswith(base):
            if k != base:  # 避免重复
                return v
    return None


def _best_sheet_name(locs, base):
    """从多个映射中选最佳sheet名：优先sheet名包含货号数字前缀的条目
    例如 92119 有 [转B版本, 92119, 转B版本]，优先选 '92119'
    """
    if not locs:
        return ''
    if len(locs) == 1:
        return locs[0]['sheet']
    num_m = re.match(r'^(\d+)', base)
    if num_m:
        prefix = num_m.group(1)
        for loc in locs:
            sname = loc['sheet']
            # sheet名以货号数字前缀开头（如"92119"、"92119明细"）
            if sname.startswith(prefix):
                return sname
    return locs[0]['sheet']


def _load_sub_schedule_map():
    """加载 sub_schedule_map.json"""
    map_path = os.path.join(os.path.dirname(__file__), 'data', 'sub_schedule_map.json')
    if not os.path.exists(map_path):
        return {}
    mtime = os.path.getmtime(map_path)
    if _sub_map_cache.get('mtime') == mtime and _sub_map_cache.get('data'):
        return _sub_map_cache['data']
    with open(map_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    _sub_map_cache['data'] = data
    _sub_map_cache['mtime'] = mtime
    return data


def lookup_schedule_info(items):
    """查找货号列表的分排期归属

    Args:
        items: 货号列表（如 ['9548UQ1-S001', '77772-S001']）
    Returns:
        {
            'matched': {货号基础: [{file, sheet}]},
            'unmatched': [货号基础列表]
        }
    """
    sub_map = _load_sub_schedule_map()
    if not sub_map:
        return {'matched': {}, 'unmatched': [s for s in items]}

    matched = {}
    unmatched = []

    for raw_item in items:
        raw = str(raw_item).strip()
        if not raw:
            continue
        # 去规格码得到基础货号
        m = re.match(r'^(.+?)(-S\d+.*)$', raw, re.I)
        base = m.group(1).upper() if m else raw.upper()

        locs = _find_in_sub_map(sub_map, base)
        if locs:
            matched[base] = locs
            continue

        unmatched.append(base)

    return {'matched': matched, 'unmatched': unmatched}
