"""扫描总排期Excel，生成 货号→中文名 直查表（item_cn_name_map.json）
用法：python scan_cn_names.py
输出：data/item_cn_name_map.json
"""
import os, sys, json, re, logging
from collections import Counter

logging.basicConfig(level=logging.INFO, format='%(message)s')

# 总排期路径
MASTER_PATH = r'Z:\业务部存档资料\曾庆博\总排期备份\新建文件夹\2025年ZURU总生产排期5.4-5.8.xlsx'
ITEM_COL = 7   # G列=货号#
CN_COL = 8     # H列=中文名
SHEET_NAME = '总排期'

def _sku_spec(s):
    """提取货号基础部分（去掉-S001等规格码），用于分组"""
    s = str(s).strip()
    if not s:
        return ''
    # 去掉规格码 -S001/-S002 等
    m = re.match(r'^(.+?)-S\d+', s, re.I)
    if m:
        return m.group(1).upper()
    return s.upper()


def _full_item(s):
    """完整货号（去空白换行，保留-S00x/-PKC等后缀），用于精确匹配"""
    s = re.sub(r'[\s\n]+', '', str(s).strip()).upper()
    return s if s else ''

def scan():
    try:
        import openpyxl
    except ImportError:
        logging.error("需要安装openpyxl: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(MASTER_PATH):
        logging.error(f"总排期文件不存在: {MASTER_PATH}")
        sys.exit(1)

    logging.info(f"正在读取总排期: {MASTER_PATH}")
    wb = openpyxl.load_workbook(MASTER_PATH, read_only=True, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if '总排期' in name and '旧' not in name:
            ws = wb[name]
            break
    if ws is None:
        logging.error(f"找不到'总排期'工作表，可用sheet: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    logging.info(f"使用工作表: {ws.title}")

    # 收集 完整货号→中文名 + 基础码→中文名 出现次数
    full_cn_counter = {}   # {完整货号: Counter({中文名: 次数})}
    base_cn_counter = {}   # {基础码: Counter({中文名: 次数})}
    total_rows = 0
    valid_rows = 0

    for row in ws.iter_rows(min_row=2, max_col=30, values_only=False):
        total_rows += 1
        cells = list(row)
        if len(cells) < CN_COL:
            continue

        item_val = cells[ITEM_COL - 1].value if len(cells) >= ITEM_COL else None
        cn_val = cells[CN_COL - 1].value if len(cells) >= CN_COL else None

        if not item_val or not cn_val:
            continue

        item_str = str(item_val).strip()
        cn_str = str(cn_val).strip()

        if not item_str or not cn_str:
            continue

        # 排除汇总行
        if any(kw in cn_str.upper() for kw in ('TOTAL', 'SUBTOTAL', '合计', '小计')):
            continue

        full = _full_item(item_str)
        spec = _sku_spec(item_str)
        if not full:
            continue

        valid_rows += 1

        # 按完整货号统计
        if full not in full_cn_counter:
            full_cn_counter[full] = Counter()
        full_cn_counter[full][cn_str] += 1

        # 按基础码统计（作为fallback）
        if spec:
            if spec not in base_cn_counter:
                base_cn_counter[spec] = Counter()
            base_cn_counter[spec][cn_str] += 1

    wb.close()

    # 构建结果：完整货号条目 + 基础码兜底条目
    result = {}

    # 1) 完整货号条目（精确匹配用）
    for full, counter in sorted(full_cn_counter.items()):
        best_cn, count = counter.most_common(1)[0]
        result[full] = {
            'cn_name': best_cn,
            'count': count,
            'total': sum(counter.values()),
            'alternatives': {k: v for k, v in counter.most_common() if k != best_cn} if len(counter) > 1 else {}
        }

    # 2) 基础码兜底条目（仅当基础码不等于任何已有完整货号时才添加）
    base_added = 0
    for spec, counter in sorted(base_cn_counter.items()):
        if spec not in result:
            best_cn, count = counter.most_common(1)[0]
            result[spec] = {
                'cn_name': best_cn,
                'count': count,
                'total': sum(counter.values()),
                'alternatives': {k: v for k, v in counter.most_common() if k != best_cn} if len(counter) > 1 else {}
            }
            base_added += 1

    # 保存
    out_dir = os.path.join(os.path.dirname(__file__), 'data')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'item_cn_name_map.json')

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    logging.info(f"\n扫描完成:")
    logging.info(f"  总行数: {total_rows}")
    logging.info(f"  有效行数: {valid_rows}")
    logging.info(f"  完整货号条目: {len(full_cn_counter)}")
    logging.info(f"  基础码兜底条目: {base_added}")
    logging.info(f"  总条目数: {len(result)}")
    logging.info(f"  保存到: {out_path}")

    # 显示有多个中文名候选的完整货号（需人工确认）
    conflicts = {k: v for k, v in result.items() if v['alternatives']}
    if conflicts:
        logging.info(f"\n有 {len(conflicts)} 个货号存在多个中文名（已选最多的）:")
        for spec, info in sorted(conflicts.items()):
            alts = ', '.join(f'"{k}"({v}次)' for k, v in info['alternatives'].items())
            logging.info(f"  {spec}: 选 \"{info['cn_name']}\"({info['count']}次) | 其他: {alts}")

    return out_path, result

if __name__ == '__main__':
    scan()
